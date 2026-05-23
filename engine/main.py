from dotenv import load_dotenv
load_dotenv()

import sys
import math
sys.stdout.reconfigure(encoding='utf-8', errors='replace')
sys.stderr.reconfigure(encoding='utf-8', errors='replace')


def make_json_safe(obj):
    """Recursively replace NaN/Inf floats with None for JSON compliance."""
    if isinstance(obj, float):
        return None if (math.isnan(obj) or math.isinf(obj)) else obj
    if isinstance(obj, dict):
        return {k: make_json_safe(v) for k, v in obj.items()}
    if isinstance(obj, list):
        return [make_json_safe(v) for v in obj]
    return obj

from fastapi.responses import StreamingResponse, JSONResponse, FileResponse
from fastapi.middleware.cors import CORSMiddleware
from fastapi.middleware.gzip import GZipMiddleware
import concurrent.futures
from pydantic import BaseModel
import polars as pl
import io
import json
import uuid
from typing import Optional, List, Union, Literal
import plotly.express as px
import plotly.graph_objects as go
import pandas as pd
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, PatternFill, Alignment
from google import genai
from contextlib import asynccontextmanager
import os
import socket
import sys
import database as db
from sqlalchemy.ext.asyncio import AsyncSession
from db_async import get_db
from auth import get_current_user
from models import User
from insight_engine import (
    ColumnClassifier, MetricComputer, BusinessRuleEngine,
    InsightNarrator, SmartChartRecommender, AnomalyDetector,
    run_insight_engine, RecommendationEngine, StrategicBriefBuilder,
    validate_dataframe, auto_clean_dataframe,
)
from session_cache import SessionCache, JobTracker
from fastapi import FastAPI, File, UploadFile, HTTPException, Form, Query, BackgroundTasks, Depends
from report_generator import ChartGenerator, PDFReportGenerator, ColumnMap, cleanup_temp_files, UnifiedReportGenerator
import report_generator

import tempfile
from pathlib import Path

SESSION_DIR = Path(tempfile.gettempdir()) / "insightstream_sessions"

# Global chart layout base to prevent duplicate titles and ensure styling
CHART_LAYOUT_BASE = {
    "title": {"text": ""},  # ← KEY FIX: empty title in chart itself
    "paper_bgcolor": "rgba(0,0,0,0)",
    "plot_bgcolor": "rgba(0,0,0,0)",
    "font": {"color": "#94a3b8"},
    "xaxis": {"gridcolor": "rgba(255,255,255,0.05)"},
    "yaxis": {"gridcolor": "rgba(255,255,255,0.05)"},
    "legend": {"bgcolor": "rgba(0,0,0,0)"},
    "margin": {"l": 60, "r": 30, "t": 20, "b": 60},
}

import re

def sanitize_insight(text: str, max_length=500) -> str:
    """Truncate text if too long, replace raw number dumps with warning."""
    if not text:
        return ""
    if not isinstance(text, str):
        text = str(text)
    # Detect raw data dumps:
    #  1) Space-separated number sequences  ("1.5 2.3 3.7 ...")
    #  2) Comma-separated number lists      ("1.5, 2.3, 3.7, ...")
    #  3) Lines that are mostly digits (>60% digit characters)
    if re.search(r'\b\d+(\.\d+)?\s+\d+(\.\d+)?\s+\d+', text):
        return "[Strategic insight unavailable – raw data suppressed]"
    if re.search(r'\d+(\.\d+)?,\s*\d+(\.\d+)?,\s*\d+(\.\d+)?,\s*\d+', text):
        return "[Strategic insight unavailable – raw data suppressed]"
    # If >60% of characters are digits/dots/commas, it's a data dump
    digit_chars = sum(1 for c in text if c in '0123456789.,')
    if len(text) > 50 and digit_chars / max(len(text), 1) > 0.6:
        return "[Strategic insight unavailable – raw data suppressed]"
    if len(text) > max_length:
        text = text[:max_length] + "…"
    return text

def is_chart_informative(values: list[float], min_variance_pct: float = 1.0) -> bool:
    """
    Returns False if all values are within min_variance_pct% of each other.
    Useful for suppressing flat-line charts that show no information.
    """
    if not values or len(values) < 2:
        return False
    
    import numpy as np
    arr = np.array(values, dtype=float)
    
    if np.all(arr == 0):
        return False
    
    mean = np.mean(arr)
    if mean == 0:
        return False
    
    # Coefficient of variation as percentage
    cv_pct = (np.std(arr) / abs(mean)) * 100
    return cv_pct >= min_variance_pct

import sqlite3
import os

async def run_migrations(db_path: str):
    """Add missing columns to existing database."""
    conn = sqlite3.connect(db_path)
    cursor = conn.cursor()
    
    # Get existing columns
    cursor.execute("PRAGMA table_info(analysis_sessions)")
    existing_cols = {row[1] for row in cursor.fetchall()}
    
    # Add currency column if missing
    if "currency" not in existing_cols:
        try:
            cursor.execute(
                "ALTER TABLE analysis_sessions "
                "ADD COLUMN currency VARCHAR(10) DEFAULT NULL"
            )
            conn.commit()
            print("[DB MIGRATION] Added 'currency' column ✓")
        except Exception as e:
            print(f"[DB MIGRATION] Error or already exists: {e}")

    # Add llm_results_json column if missing
    if "llm_results_json" not in existing_cols:
        try:
            cursor.execute(
                "ALTER TABLE analysis_sessions "
                "ADD COLUMN llm_results_json TEXT DEFAULT NULL"
            )
            conn.commit()
            print("[DB MIGRATION] Added 'llm_results_json' column ✓")
        except Exception as e:
            print(f"[DB MIGRATION] Error or already exists: {e}")

    conn.close()

@asynccontextmanager
async def lifespan(app: FastAPI):
    """Create session directory and initialize database on startup."""
    try:
        SESSION_DIR.mkdir(parents=True, exist_ok=True)
        print(f"Session directory created: {SESSION_DIR}")
    except Exception as e:
        print(f"Warning: Could not create session directory: {e}")
    # Initialize SQLite database
    db.init_db()
    print("Database initialized.")
    
    # Run migrations
    db_path = str(Path(os.environ.get("INSIGHTSTREAM_DATA_DIR", Path(__file__).parent / "data")) / "insightstream.db")
    await run_migrations(db_path)
    
    yield

app = FastAPI(title="Virtual Data Scientist Engine", lifespan=lifespan)

from routers.auth import router as auth_router
from routers.analyze import router as analyze_router
from routers.sessions import router as sessions_router
from routers.chat import router as chat_router
app.include_router(auth_router)
app.include_router(analyze_router)
app.include_router(sessions_router)
app.include_router(chat_router)

# Allow CORS for Next.js frontend (localhost + production)
app.add_middleware(
    CORSMiddleware,
    allow_origins=[
        "http://localhost:3000",
        "https://insightstream-alpha.vercel.app",
        "https://*.vercel.app",
    ],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# GZip compress responses > 1KB (charts payload: 500KB → 80KB)
app.add_middleware(GZipMiddleware, minimum_size=1000)

# ============== PERFORMANCE: CACHE + BACKGROUND JOBS ==============
_cache = SessionCache()
_jobs  = JobTracker()
_thread_pool = concurrent.futures.ThreadPoolExecutor(max_workers=2)

# ============== FILE-BASED SESSION STORAGE ==============
# Sessions are stored in /tmp/sessions/ as JSON (metadata) + Parquet (dataframe)
# This survives within a container's lifetime, unlike in-memory dicts

# Configure Gemini
GENAI_API_KEY = os.getenv("GEMINI_API_KEY")
client = None
if GENAI_API_KEY:
    client = genai.Client(api_key=GENAI_API_KEY)

def save_session(session_id: str, filename: str, df: pl.DataFrame) -> None:
    """Save session to disk."""
    # Ensure base directory exists (fallback in case startup event didn't run)
    SESSION_DIR.mkdir(parents=True, exist_ok=True)
    
    session_path = SESSION_DIR / session_id
    session_path.mkdir(parents=True, exist_ok=True)
    
    # Save metadata
    metadata = {"filename": filename}
    with open(session_path / "metadata.json", "w") as f:
        json.dump(metadata, f)
    
    # Save DataFrame as Parquet (fast and efficient)
    df.write_parquet(session_path / "data.parquet")

def load_session(session_id: str) -> tuple[str, pl.DataFrame]:
    """Load session from disk. Checks legacy parquet store first, then uploads dir."""
    # ── Legacy path (old file-based sessions) ─────────────────────────
    session_path = SESSION_DIR / session_id
    if session_path.exists():
        data_file = session_path / "data.parquet"
        metadata_file = session_path / "metadata.json"
        if not data_file.exists():
            raise FileNotFoundError(f"Session data file not found for {session_id}")
        if not metadata_file.exists():
            raise FileNotFoundError(f"Session metadata not found for {session_id}")
        try:
            with open(metadata_file, "r") as f:
                metadata = json.load(f)
        except Exception as e:
            raise ValueError(f"Failed to load session metadata: {str(e)}")
        try:
            df = pl.read_parquet(data_file)
        except Exception as e:
            raise ValueError(f"Failed to load session data: {str(e)}")
        return metadata["filename"], df

    # ── New path (DB-backed sessions with uploaded file on disk) ──────
    import glob as _glob
    upload_dir = Path(__file__).parent / "data" / "uploads"
    matches = _glob.glob(str(upload_dir / f"session_{session_id}_*"))
    if matches:
        filepath = Path(matches[0])
        lower = filepath.name.lower()
        try:
            if lower.endswith(".xlsx") or lower.endswith(".xls"):
                df_pd = pd.read_excel(filepath)
            else:
                try:
                    df_pd = pd.read_csv(filepath, encoding="utf-8")
                except UnicodeDecodeError:
                    df_pd = pd.read_csv(filepath, encoding="latin-1")
        except Exception as e:
            raise ValueError(f"Failed to load uploaded file for session {session_id}: {e}")

        # Normalize object/string columns to str before Polars conversion —
        # mixed-type columns (e.g. int + str) cause ArrowTypeError.
        for col in df_pd.select_dtypes(include=["object", "string"]).columns:
            df_pd[col] = df_pd[col].astype(str)

        try:
            return filepath.name, pl.from_pandas(df_pd)
        except Exception as e:
            # Fall back to pandas if Polars conversion still fails
            print(f"[load_session] Polars conversion failed: {e} — using pandas")
            return filepath.name, df_pd

    raise FileNotFoundError(f"Session {session_id} not found")

def session_exists(session_id: str) -> bool:
    """Check if a session exists in either the legacy parquet store or the uploads dir."""
    if (SESSION_DIR / session_id / "data.parquet").exists():
        return True
    import glob as _glob
    upload_dir = Path(__file__).parent / "data" / "uploads"
    return bool(_glob.glob(str(upload_dir / f"session_{session_id}_*")))


# ── DB-backed health-check endpoint ───────────────────────────────────────────
@app.get("/health-check/{session_id}")
async def health_check_endpoint(
    session_id: int,
    current_user: User = Depends(get_current_user),
    db_session: AsyncSession = Depends(get_db),
):
    from services.session_service import get_session_detail
    session = await get_session_detail(db_session, session_id, current_user.id)
    if not session:
        raise HTTPException(404, "Session not found")

    try:
        _, df_raw = load_session(str(session_id))
        df = df_raw.to_pandas() if hasattr(df_raw, "to_pandas") else df_raw
    except FileNotFoundError:
        raise HTTPException(404, "Session file not found on disk")

    total_rows = len(df)
    total_cols = len(df.columns)
    duplicates = int(df.duplicated().sum())
    issues = []

    for col in df.columns:
        missing = int(df[col].isna().sum())
        if missing > 0:
            pct = round(missing / total_rows * 100, 2)
            issues.append({
                "column": col,
                "issue_type": "missing",
                "severity": "high" if pct > 10 else "low",
                "description": f"{missing} missing values ({pct}%)",
                "count": missing,
                "percentage": pct,
                "suggested_action": "drop_column" if pct > 70 else "impute_median",
            })

    for col in df.select_dtypes(include="number").columns:
        q1 = df[col].quantile(0.25)
        q3 = df[col].quantile(0.75)
        iqr = q3 - q1
        if iqr == 0:
            continue
        outlier_count = int(((df[col] < q1 - 1.5 * iqr) | (df[col] > q3 + 1.5 * iqr)).sum())
        if outlier_count > 0:
            pct = round(outlier_count / total_rows * 100, 2)
            issues.append({
                "column": col,
                "issue_type": "outlier",
                "severity": "high" if pct > 5 else "low",
                "description": f"{outlier_count} outliers detected ({pct}%)",
                "count": outlier_count,
                "percentage": pct,
            })

    if duplicates > 0:
        pct = round(duplicates / total_rows * 100, 2)
        issues.append({
            "column": "_all_",
            "issue_type": "duplicate",
            "severity": "high" if pct > 5 else "medium",
            "description": f"{duplicates} duplicate rows ({pct}%)",
            "count": duplicates,
            "percentage": pct,
            "suggested_action": "drop_duplicates",
        })

    for col in df.columns:
        if df[col].nunique() <= 1:
            issues.append({
                "column": col,
                "issue_type": "constant",
                "severity": "medium",
                "description": f"Column has only {df[col].nunique()} unique value — no variance",
                "count": total_rows,
                "percentage": 100.0,
                "suggested_action": "drop_column",
            })

    for col in df.select_dtypes(include="object").columns:
        n_unique = df[col].nunique()
        pct_unique = n_unique / total_rows if total_rows > 0 else 0
        if pct_unique > 0.95 and total_rows > 100:
            issues.append({
                "column": col,
                "issue_type": "high_cardinality",
                "severity": "low",
                "description": f"{n_unique} unique values ({round(pct_unique * 100, 1)}%) — likely an ID column",
                "count": n_unique,
                "percentage": round(pct_unique * 100, 2),
                "suggested_action": "review_column",
            })

    high_issues = [i for i in issues if i.get("severity") == "high"]
    missing_issues = [i for i in issues if i.get("issue_type") == "missing"]
    total_missing_pct = sum(i.get("percentage", 0) for i in missing_issues)

    if len(issues) == 0 and duplicates == 0:
        score = "A"
    elif len(high_issues) == 0:
        score = "B"
    elif total_missing_pct > 50 or duplicates > total_rows * 0.5:
        score = "D"
    else:
        score = "C"

    return {
        "session_id": str(session_id),
        "quality_score": score,
        "row_count": total_rows,
        "column_count": total_cols,
        "duplicate_rows": duplicates,
        "issues": issues,
    }


# ── DB-backed EDA endpoint ─────────────────────────────────────────────────────
@app.get("/eda/{session_id}")
async def eda_endpoint(
    session_id: int,
    current_user: User = Depends(get_current_user),
    db_session: AsyncSession = Depends(get_db),
):
    from services.session_service import get_session_detail
    session = await get_session_detail(db_session, session_id, current_user.id)
    if not session:
        raise HTTPException(404, "Session not found")

    try:
        _, df_raw = load_session(str(session_id))
        df = df_raw.to_pandas() if hasattr(df_raw, "to_pandas") else df_raw
    except FileNotFoundError:
        raise HTTPException(404, "Session file not found on disk")

    numeric_cols, categorical_cols, date_cols, identifier_cols, binary_cols = [], [], [], [], []
    column_stats = []
    warnings = []

    for col in df.columns:
        dtype_str = str(df[col].dtype)
        n_unique = int(df[col].nunique())
        n_missing = int(df[col].isna().sum())

        if pd.api.types.is_datetime64_any_dtype(df[col]):
            role = "temporal"
            date_cols.append(col)
        elif pd.api.types.is_numeric_dtype(df[col]):
            if n_unique == 2:
                role = "binary"
                binary_cols.append(col)
            elif n_unique / max(len(df), 1) > 0.95 and n_unique > 50:
                role = "identifier"
                identifier_cols.append(col)
            else:
                role = "numerical"
                numeric_cols.append(col)
        elif df[col].dtype == object or pd.api.types.is_string_dtype(df[col]):
            if n_unique == 2:
                role = "binary"
                binary_cols.append(col)
            elif n_unique > 5:
                # Try date parsing before falling back to categorical/identifier
                try:
                    sample = df[col].dropna().head(200)
                    # Try dayfirst=True first (handles DD/MM/YYYY like "1/8/2023 11:13")
                    parsed = pd.to_datetime(sample, format="mixed", dayfirst=True, errors="coerce")
                    if parsed.notna().mean() <= 0.7:
                        # Fall back to MM/DD/YYYY
                        parsed = pd.to_datetime(sample, format="mixed", errors="coerce")
                    if parsed.notna().mean() > 0.7:
                        role = "temporal"
                        date_cols.append(col)
                    elif n_unique <= 100:
                        role = "categorical"
                        categorical_cols.append(col)
                    else:
                        role = "identifier"
                        identifier_cols.append(col)
                except Exception:
                    if n_unique <= 100:
                        role = "categorical"
                        categorical_cols.append(col)
                    else:
                        role = "identifier"
                        identifier_cols.append(col)
            elif n_unique <= 100:
                role = "categorical"
                categorical_cols.append(col)
            else:
                role = "identifier"
                identifier_cols.append(col)
        else:
            role = "categorical"
            categorical_cols.append(col)

        stat: dict = {
            "column": col,
            "dtype": dtype_str,
            "role": role,
            "unique_count": n_unique,
            "mean": None,
            "median": None,
            "std": None,
            "min_val": None,
            "max_val": None,
            "top_values": [],
        }
        if pd.api.types.is_numeric_dtype(df[col]):
            stat["mean"] = round(float(df[col].mean()), 2) if not df[col].isna().all() else None
            stat["median"] = round(float(df[col].median()), 2) if not df[col].isna().all() else None
            stat["std"] = round(float(df[col].std()), 2) if not df[col].isna().all() else None
            stat["min_val"] = round(float(df[col].min()), 2) if not df[col].isna().all() else None
            stat["max_val"] = round(float(df[col].max()), 2) if not df[col].isna().all() else None
        else:
            top = df[col].value_counts().head(5)
            stat["top_values"] = [{"value": str(v), "count": int(c)} for v, c in top.items()]

        if n_missing > 0:
            warnings.append(f"{col}: {n_missing} missing values ({round(n_missing/len(df)*100,1)}%)")

        column_stats.append(stat)

    correlation_matrix: dict = {}
    num_df = df[numeric_cols] if numeric_cols else pd.DataFrame()
    if len(numeric_cols) >= 2:
        try:
            corr = num_df.corr().round(3)
            correlation_matrix = corr.to_dict()
        except Exception:
            pass

    insights: list[str] = []
    if numeric_cols:
        insights.append(f"Dataset has {len(numeric_cols)} numeric columns available for analysis.")
    if categorical_cols:
        insights.append(f"{len(categorical_cols)} categorical columns can be used for segmentation.")
    if date_cols:
        insights.append(f"Time-series analysis is possible using {', '.join(date_cols)}.")
    if identifier_cols:
        insights.append(f"Columns {', '.join(identifier_cols)} appear to be identifiers and are excluded from model training.")

    return make_json_safe({
        "session_id": str(session_id),
        "numeric_columns": numeric_cols,
        "categorical_columns": categorical_cols,
        "date_columns": date_cols,
        "identifier_columns": identifier_cols,
        "binary_columns": binary_cols,
        "column_stats": column_stats,
        "correlation_matrix": correlation_matrix,
        "insights": insights,
        "warnings": warnings,
    })


# ── Outlier rows endpoint ──────────────────────────────────────────────────────
@app.get("/eda/{session_id}/outliers/{column}")
async def get_outlier_rows(
    session_id: int,
    column: str,
    page: int = 1,
    page_size: int = 50,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    from services.session_service import get_session_detail
    session = await get_session_detail(db, session_id, current_user.id)
    if not session:
        raise HTTPException(404, "Session not found")

    try:
        _, df_raw = load_session(str(session_id))
        df = df_raw.to_pandas() if hasattr(df_raw, "to_pandas") else df_raw
    except FileNotFoundError:
        raise HTTPException(404, "Session file not found on disk")

    if column not in df.columns:
        raise HTTPException(404, f"Column '{column}' not found")

    if not pd.api.types.is_numeric_dtype(df[column]):
        raise HTTPException(400, f"Column '{column}' is not numeric")

    col_data = df[column].dropna()
    Q1 = float(col_data.quantile(0.25))
    Q3 = float(col_data.quantile(0.75))
    IQR = Q3 - Q1
    lower_bound = Q1 - 1.5 * IQR
    upper_bound = Q3 + 1.5 * IQR

    mask = (df[column] < lower_bound) | (df[column] > upper_bound)
    outlier_df = df[mask].copy()
    outlier_df["_outlier_value"] = outlier_df[column]
    outlier_df["_outlier_direction"] = outlier_df[column].apply(
        lambda v: "high" if v > upper_bound else "low"
    )

    total_outliers = len(outlier_df)
    pct = round(total_outliers / max(len(df), 1) * 100, 2)
    total_pages = max(1, (total_outliers + page_size - 1) // page_size)

    start = (page - 1) * page_size
    page_df = outlier_df.iloc[start : start + page_size]

    return make_json_safe({
        "column": column,
        "total_outliers": total_outliers,
        "pct": pct,
        "lower_bound": round(lower_bound, 4),
        "upper_bound": round(upper_bound, 4),
        "Q1": round(Q1, 4),
        "Q3": round(Q3, 4),
        "page": page,
        "page_size": page_size,
        "total_pages": total_pages,
        "rows": page_df.to_dict(orient="records"),
        "columns": list(page_df.columns),
    })


# ── Missing value rows endpoint ────────────────────────────────────────────────
@app.get("/eda/{session_id}/missing/{column}")
async def get_missing_rows(
    session_id: int,
    column: str,
    page: int = 1,
    page_size: int = 50,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    from services.session_service import get_session_detail
    session = await get_session_detail(db, session_id, current_user.id)
    if not session:
        raise HTTPException(404, "Session not found")
    try:
        _, df_raw = load_session(str(session_id))
        df = df_raw.to_pandas() if hasattr(df_raw, "to_pandas") else df_raw
    except FileNotFoundError:
        raise HTTPException(404, "Session file not found on disk")

    if column not in df.columns:
        raise HTTPException(404, f"Column '{column}' not found")

    missing_df = df[df[column].isna()].copy()
    missing_df["_row_index"] = missing_df.index

    total = len(missing_df)
    total_pages = max(1, (total + page_size - 1) // page_size)
    start = (page - 1) * page_size
    page_df = missing_df.iloc[start : start + page_size]

    return make_json_safe({
        "column": column,
        "total_missing": total,
        "pct": round(total / max(len(df), 1) * 100, 2),
        "page": page,
        "page_size": page_size,
        "total_pages": total_pages,
        "rows": page_df.to_dict(orient="records"),
        "columns": list(df.columns),
    })


# ── Duplicate rows endpoint ────────────────────────────────────────────────────
@app.get("/eda/{session_id}/duplicates")
async def get_duplicate_rows(
    session_id: int,
    page: int = 1,
    page_size: int = 50,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    from services.session_service import get_session_detail
    session = await get_session_detail(db, session_id, current_user.id)
    if not session:
        raise HTTPException(404, "Session not found")
    try:
        _, df_raw = load_session(str(session_id))
        df = df_raw.to_pandas() if hasattr(df_raw, "to_pandas") else df_raw
    except FileNotFoundError:
        raise HTTPException(404, "Session file not found on disk")

    dup_mask = df.duplicated(keep=False)
    dup_df = df[dup_mask].copy()
    dup_df["_row_index"] = dup_df.index
    dup_df["_dup_group"] = df[dup_mask].groupby(list(df.columns)).ngroup().values

    total = len(dup_df)
    total_pages = max(1, (total + page_size - 1) // page_size)
    start = (page - 1) * page_size
    page_df = dup_df.iloc[start : start + page_size]

    return make_json_safe({
        "total_duplicates": total,
        "pct": round(total / max(len(df), 1) * 100, 2),
        "page": page,
        "page_size": page_size,
        "total_pages": total_pages,
        "rows": page_df.to_dict(orient="records"),
        "columns": list(df.columns),
    })


class ColumnInfo(BaseModel):
    name: str
    dtype: str
    missing_count: int
    missing_pct: float
    unique_count: int

class DatasetSchema(BaseModel):
    session_id: Optional[str] = None
    filename: str
    row_count: Optional[int] = None
    column_count: Optional[int] = None
    columns: Optional[list[ColumnInfo]] = None
    preview: Optional[list[dict]] = None  # First 10 rows
    sheets: Optional[List[str]] = None    # For multi-sheet Excel files
    requires_selection: bool = False      # Flag to prompt user for sheet selection
    quality_report: Optional[dict] = None # Data quality audit result

@app.get("/")
def read_root():
    return {"status": "online", "message": "Virtual Data Scientist Engine is running"}

@app.get("/health")
def health_check():
    return {"status": "ok"}

# ─── Project CRUD Endpoints ─────────────────────────────────────

class ProjectRename(BaseModel):
    name: str

@app.get("/projects")
def list_all_projects():
    """List all saved projects."""
    return db.list_projects()

@app.get("/projects/{project_id}")
def get_project_detail(project_id: str):
    """Get project details and history."""
    project = db.get_project(project_id)
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")
    history = db.get_history(project_id)
    return {**project, "history": history}

@app.put("/projects/{project_id}")
def rename_project(project_id: str, body: ProjectRename):
    """Rename a project."""
    if not db.update_project(project_id, name=body.name):
        raise HTTPException(status_code=404, detail="Project not found")
    return {"status": "ok", "name": body.name}

@app.delete("/projects/{project_id}")
def delete_project_endpoint(project_id: str):
    """Delete a project and its data."""
    if not db.delete_project(project_id):
        raise HTTPException(status_code=404, detail="Project not found")
    return {"status": "ok", "deleted": project_id}

# ─── Dashboard Endpoints ────────────────────────────────────────

class DashboardSave(BaseModel):
    layout: list
    pinned_chart_ids: list[str]
    text_blocks: Optional[list[dict]] = None

@app.post("/dashboard/{project_id}")
def save_dashboard_endpoint(project_id: str, body: DashboardSave):
    """Save dashboard layout for a project."""
    db.save_dashboard(project_id, body.layout, body.pinned_chart_ids, body.text_blocks)
    return {"status": "ok"}

class DashboardExportRequest(BaseModel):
    title: str
    template: str = "modern"
    project_name: str = "InsightStream"
    kpis: dict
    charts: list[dict] # {id, title, image_base64, error, insight}
    ai_summary: Optional[str] = ""
    insights: Optional[list] = []
    recommendations: Optional[list] = []
    text_blocks: Optional[list[dict]] = []

@app.post("/export-dashboard-pdf/{session_id}")
async def export_dashboard_pdf(
    session_id: str,
    body: DashboardExportRequest,
    background_tasks: BackgroundTasks,
    db: AsyncSession = Depends(get_db)
):
    """
    Unified Pixel-Perfect Export (Multi-page).
    Receives canonical dashboard assets and assembles a professional PDF.
    """
    import logging
    logger = logging.getLogger(__name__)
    logger.info(f"🚀 [PIPELINE] New Pixel-Perfect Export hit for session: {session_id}")
    logger.info(f"📦 [PAYLOAD] Received {len(body.charts)} charts, {len(body.insights)} insights.")

    # ── Resolve user_id (optional — shared links have no auth) ───────────
    user_id = None
    try:
        from auth import get_current_user_optional
        _cu = await get_current_user_optional(db)
        user_id = _cu.id if _cu else None
    except Exception:
        pass

    # ── Load currency from session ────────────────────────────────────────
    _currency_code = None
    try:
        from services.session_service import get_session_detail as _gsd_cur
        if user_id:
            _sess_cur = await _gsd_cur(db, int(session_id), user_id)
            if _sess_cur:
                _currency_code = getattr(_sess_cur, "currency", None)
                print(f"[PDF EXPORT] Currency override: {_currency_code}")
    except Exception as _ce:
        print(f"[PDF EXPORT] Currency load failed: {_ce}")

    # ── Helper: plotly_json string → base64 PNG ───────────────────────────
    def _plotly_json_to_base64(plotly_json_str: str):
        """Deserialize stored Plotly JSON and render to PNG base64.
        Strips layout.template first — it may contain trace types (e.g. heatmapgl)
        removed in newer Plotly versions that cause from_json() to crash.
        """
        try:
            import plotly.io as _pio
            import base64 as _b64
            import json as _json
            raw = _json.loads(plotly_json_str)
            raw.get("layout", {}).pop("template", None)
            _fig = _pio.from_json(_json.dumps(raw))
            _img = _pio.to_image(_fig, format="png", width=800, height=500)
            return _b64.b64encode(_img).decode("utf-8")
        except Exception as _err:
            print(f"[EXPORT] Chart conversion failed: {_err}")
            return None

    try:
        out_name = f"Report_{session_id}_{uuid.uuid4().hex[:6]}.pdf"
        out_path = Path(tempfile.gettempdir()) / out_name

        # ── Load DF from disk ─────────────────────────────────────────────
        try:
            filename, df = load_session(session_id)
            from insight_engine import detect_domain
            domain_id = detect_domain(df)
        except Exception as _le:
            logger.warning(f"Could not load session data for PDF: {_le}")
            df = None
            domain_id = "general"

        # ── UNCONDITIONAL LLM OVERRIDE ────────────────────────────────────
        # Check DB for stored LLM results regardless of auth state.
        # Uses raw SQLite so it works even when user_id is None.
        _llm_override = False
        _insights_to_use = None
        _structured_recs = None
        _llm_domain_id = None

        try:
            import sqlite3 as _sq
            import json as _jmod
            from pathlib import Path as _P
            _db_path = _P(__file__).parent / "data" / "insightstream.db"
            _conn = _sq.connect(str(_db_path))
            _row = _conn.execute(
                "SELECT llm_results_json FROM analysis_sessions WHERE id = ?",
                (int(session_id),)
            ).fetchone()
            _conn.close()

            if _row and _row[0]:
                _llm_data = _jmod.loads(_row[0])
                _stored_charts = _llm_data.get("charts", [])

                if _stored_charts:
                    _converted = []
                    for _ch in _stored_charts:
                        _b64 = _plotly_json_to_base64(_ch.get("plotly_json", ""))
                        if _b64:
                            _converted.append({
                                "id":           _ch.get("id", "llm_chart"),
                                "title":        _ch.get("title", ""),
                                "image_base64": _b64,
                                "insight":      "",
                            })
                            print(f"[EXPORT] Converted chart: {_ch.get('title')!r}")
                        else:
                            print(f"[EXPORT] Skipped chart: {_ch.get('title')!r}")

                    if _converted:
                        body.charts = _converted
                        _llm_override = True
                        _insights_to_use = _llm_data.get("insights", [])
                        _structured_recs = [
                            {
                                "action":    r.get("text", r.get("action", "")),
                                "timeframe": r.get("timeframe", "Next 30 days"),
                                "owner":     r.get("owner", "Strategy team"),
                                "impact":    r.get("impact", "Important"),
                            }
                            for r in _llm_data.get("recommendations", [])
                        ]
                        _llm_domain_id = _llm_data.get("domain", "general").lower()
                        domain_id = _llm_domain_id
                        print(f"[EXPORT] _llm_override=True, {len(_converted)} LLM charts ready")
        except Exception as _oe:
            print(f"[EXPORT] LLM override check failed: {_oe}")
        # ── END UNCONDITIONAL LLM OVERRIDE ───────────────────────────────

        logger.info(f"🛠 [GENERATOR] domain={domain_id}, llm_override={_llm_override}")

        # ── Sanitize payload ──────────────────────────────────────────────
        clean_ai_summary = sanitize_insight(body.ai_summary, max_length=1000)
        clean_insights = [
            ins if isinstance(ins, dict) else sanitize_insight(ins, max_length=500)
            for ins in body.insights
        ]
        clean_charts = []
        for ch in body.charts:
            ch_copy = ch.copy()
            if "insight" in ch_copy:
                ch_copy["insight"] = sanitize_insight(ch_copy["insight"], max_length=400)
            clean_charts.append(ch_copy)

        # ── Chart fallback (rule engine) — ONLY for non-LLM sessions ─────
        if not _llm_override:
            needs_plotly = any(
                not ch.get("image_base64") and not ch.get("plotly_json")
                for ch in clean_charts
            )
            if needs_plotly and clean_charts:
                print("[PDF Export] plotly_json absent — regenerating from viz engine")
                try:
                    viz_resp = _internal_gen_viz(session_id=session_id, max_charts=12)
                    viz_by_id    = {c.chart_id: c for c in viz_resp.charts}
                    viz_by_title = {c.title: c    for c in viz_resp.charts}
                    for ch in clean_charts:
                        if not ch.get("image_base64") and not ch.get("plotly_json"):
                            match = (viz_by_id.get(ch.get("id", "")) or
                                     viz_by_title.get(ch.get("title", "")))
                            if match:
                                ch["plotly_json"] = match.plotly_json
                except Exception as _ve:
                    print(f"[PDF Export] Viz fallback failed: {_ve}")

        # ── Insights / recommendations ────────────────────────────────────
        def _norm_impact(raw: str) -> str:
            s = raw.lower()
            if "critical" in s: return "critical"
            if "important" in s or "high" in s: return "important"
            return "minor"

        sports_meta = {}
        if _llm_override:
            # LLM path — use stored insights and recs
            final_insights = _insights_to_use or []
            final_recs     = _structured_recs or []
        else:
            # Rule engine path — pull from in-memory cache
            insight_result = _cache.get(session_id, "insight_result")
            _rule_insights = []
            final_recs     = []
            if insight_result and isinstance(insight_result, dict):
                _rule_insights = insight_result.get("strategic_brief", [])
                sports_meta    = insight_result.get("sports_meta", {}) or {}
                final_recs     = [
                    {**r, "impact": _norm_impact(r.get("impact", "minor"))}
                    for r in insight_result.get("recommendations", [])
                    if isinstance(r, dict)
                ]
            if not final_recs and body.recommendations:
                final_recs = [
                    r if isinstance(r, dict) else {"action": str(r)}
                    for r in body.recommendations
                ]
            final_insights = _rule_insights if _rule_insights else clean_insights

        # ── Filter hallucinated recommendations before PDF generation ─────
        # Always run — the frontend payload may contain stale/bad recs from
        # a previous session regardless of whether LLM override is active.
        if df is not None:
            try:
                from analyzer import _filter_recommendations
                _df_pd = df.to_pandas() if hasattr(df, "to_pandas") else df
                _before = len(final_recs)
                _filtered = _filter_recommendations(final_recs, _df_pd.columns.tolist())
                if len(_filtered) < _before:
                    print(f"[EXPORT] Filtered {_before - len(_filtered)} bad recs from payload")
                else:
                    print(f"[EXPORT] Rec filter: all {len(_filtered)} recs are clean")

                # Safety net: if filter removed everything, fall back to stored DB recs
                if len(_filtered) == 0 and _llm_override:
                    _db_recs = _llm_override.get("recommendations", [])
                    if _db_recs:
                        print(f"[EXPORT] All payload recs filtered — using {len(_db_recs)} stored DB recs")
                        final_recs = _db_recs
                    else:
                        final_recs = _filtered
                else:
                    final_recs = _filtered
            except Exception as _fe:
                print(f"[EXPORT] Rec filter failed: {_fe}")

        print(f"[EXPORT] Final: {len(clean_charts)} charts, "
              f"{len(final_insights)} insights, {len(final_recs)} recs, "
              f"domain={domain_id!r}")

        gen = UnifiedReportGenerator()
        gen.build_from_assets(
            output_path=str(out_path),
            charts=clean_charts,
            kpis=body.kpis,
            ai_summary=clean_ai_summary,
            insights=final_insights,
            recommendations=final_recs,
            text_blocks=body.text_blocks,
            title=body.title,
            project_name=body.project_name,
            template=body.template,
            session_id=session_id,
            df=df,
            domain_id=domain_id,
            currency_override=_currency_code,
            sports_meta=sports_meta,
        )
        logger.info(f"✅ [SUCCESS] PDF generated at: {out_path}")

        img_dir = Path(tempfile.gettempdir()) / f"insightstream_export_{session_id}"

        def _cleanup():
            try:
                os.remove(str(out_path))
            except Exception:
                pass
            try:
                cleanup_temp_files("", str(img_dir))
            except Exception:
                pass

        background_tasks.add_task(_cleanup)

        return FileResponse(
            path=str(out_path),
            filename=f"{body.title.replace(' ', '_')}.pdf",
            media_type="application/pdf",
        )
    except Exception as e:
        logger.error(f"❌ [FAILURE] PDF Export failed: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/dashboard/{project_id}")
def load_dashboard_endpoint(project_id: str):
    """Load dashboard layout for a project."""
    dashboard = db.load_dashboard(project_id)
    if not dashboard:
        return {"project_id": project_id, "layout": [], "pinned_chart_ids": [], "text_blocks": []}
    return dashboard

@app.get("/share/dashboard/{project_id}")
def load_shared_dashboard(project_id: str):
    """Public read-only payload for shared dashboard links."""
    dashboard = db.load_dashboard(project_id)
    if not dashboard:
        return {"project_id": project_id, "layout": [], "pinned_chart_ids": [], "text_blocks": [], "charts": []}

    try:
        viz = generate_visualizations(project_id, max_charts=12)
        charts = viz.get("charts", [])
    except Exception:
        charts = []

    return {
        "project_id": project_id,
        "layout": dashboard.get("layout", []),
        "pinned_chart_ids": dashboard.get("pinned_chart_ids", []),
        "text_blocks": dashboard.get("text_blocks", []),
        "charts": charts,
    }

# ─── Upload (creates project) ───────────────────────────────────

@app.post("/upload", response_model=DatasetSchema)
async def upload_dataset(
    file: UploadFile = File(...),
    sheet_name: Optional[str] = Form(None)
):
    print(f"=== UPLOAD DEBUG ===")
    print(f"Filename: {file.filename}")
    print(f"Content-Type: {file.content_type}")
    print(f"File size: {file.size if hasattr(file, 'size') else 'unknown'}")
    
    try:
        contents = await file.read()
        print(f"Bytes read: {len(contents)}")
        
        filename = file.filename.lower().strip()
        
        if filename.endswith('.csv'):
            print("Attempting CSV parse...")
            # Try multiple encodings
            df = None
            errors = []
            for encoding in ['utf-8', 'utf-8-sig', 'latin-1', 'cp1252']:
                try:
                    import io
                    df = pl.read_csv(
                        io.BytesIO(contents),
                        encoding=encoding,
                        ignore_errors=True
                    )
                    print(f"CSV parsed with encoding: {encoding}")
                    print(f"Shape: {df.shape}")
                    break
                except Exception as e:
                    errors.append(f"{encoding}: {str(e)}")
                    continue
            
            if df is None:
                print(f"All CSV encodings failed: {errors}")
                raise ValueError(f"CSV parse failed: {errors}")
                
        elif filename.endswith(('.xlsx', '.xls')):
            print("Attempting Excel parse...")
            try:
                import pandas as pd
                import io
                excel_file = pd.ExcelFile(io.BytesIO(contents), engine='openpyxl')
                sheet_names = excel_file.sheet_names
                
                if len(sheet_names) > 1 and not sheet_name:
                    return DatasetSchema(
                        filename=file.filename,
                        sheets=sheet_names,
                        requires_selection=True
                    )
                
                target_sheet = sheet_name if sheet_name else 0
                
                # Read everything as string first — prevents mixed-type crash
                pdf = pd.read_excel(
                    excel_file, 
                    sheet_name=target_sheet,
                    dtype=str,          # ← KEY FIX: read ALL columns as string
                    keep_default_na=False  # ← don't convert empty cells yet
                )
                
                # Replace empty strings with NaN
                pdf = pdf.replace('', float('nan'))
                
                # Now intelligently coerce each column
                # First, force-convert known numeric columns (insurance domain)
                FORCE_NUMERIC_PATTERNS = [
                    "amt", "amount", "payment", "price", "cost", "revenue", "sales",
                    "vintage", "tenure", "age", "years", "commission", "premium",
                    "salary", "wage", "fee", "charge", "balance", "value"
                ]
                
                for col in pdf.columns:
                    col_lower = col.lower().replace(" ", "").replace("_", "")
                    
                    # Force numeric conversion for known numeric patterns
                    force_numeric = any(pattern in col_lower for pattern in FORCE_NUMERIC_PATTERNS)
                    
                    # Try integer first
                    try:
                        converted = pd.to_numeric(pdf[col], errors='coerce')
                        non_null_orig = pdf[col].notna().sum()
                        non_null_conv = converted.notna().sum()
                        
                        # Use 50% threshold normally, but force if pattern matches
                        threshold = 0.5 if not force_numeric else 0.1  # Lower threshold for known numeric columns
                        
                        if non_null_orig > 0 and (non_null_conv / non_null_orig) >= threshold:
                            pdf[col] = converted  # Use numeric — bad rows become NaN
                            if force_numeric:
                                print(f"[FORCE NUMERIC] {col} → {non_null_conv}/{non_null_orig} values converted")
                        else:
                            if force_numeric:
                                print(f"[FORCE NUMERIC FAILED] {col} → only {non_null_conv}/{non_null_orig} values convertible (threshold={threshold})")
                        # else: leave as string for validator to flag
                    except Exception as e:
                        if force_numeric:
                            print(f"[FORCE NUMERIC ERROR] {col} → {type(e).__name__}: {str(e)}")
                        pass  # Leave as string
                
                # Safe conversion to Polars — all columns are now either float or str
                df = pl.from_pandas(pdf)
                print(f"Excel parsed. Shape: {df.shape}")
                
                # Debug: Show numeric columns detected
                numeric_cols = [c for c in df.columns if df[c].dtype in [pl.Int64, pl.Int32, pl.Float64, pl.Float32]]
                print(f"[NUMERIC COLS DETECTED] {len(numeric_cols)} columns: {numeric_cols[:10]}")  # Show first 10
                
                # Debug: Show all column names and types for empty numeric columns
                for col in df.columns:
                    if col in ["MINPAYMENTAMT", "Vintage"]:
                        non_null = df[col].drop_nulls().len()
                        print(f"[COLUMN DEBUG] {col}: dtype={df[col].dtype}, non_null_count={non_null}, sample={df[col].head(5).to_list()}")
            except Exception as e:
                print(f"Excel parse failed: {type(e).__name__}: {str(e)}")
                # Install check
                try:
                    import openpyxl
                    print(f"openpyxl version: {openpyxl.__version__}")
                except ImportError:
                    print("openpyxl NOT INSTALLED — run: pip install openpyxl")
                raise
        else:
            raise ValueError(f"Unknown file type: {filename}")
            
        print(f"=== UPLOAD SUCCESS: {df.shape} ===")

        # ── Data Quality Audit ────────────────────────────────────────────
        quality_report = validate_dataframe(df)
        print(f"[quality] critical={quality_report['summary']['critical']}  medium={quality_report['summary']['medium']}")
        print(f"[quality] DataFrame shape after validation: {df.shape}")

        if quality_report["summary"]["critical"] > 0:
            print(f"[quality] CRITICAL ISSUES FOUND - returning early without session")
            print(f"[quality] Issues: {quality_report.get('issues', [])}")
            return DatasetSchema(
                filename=file.filename,
                row_count=len(df),
                column_count=len(df.columns),
                quality_report=quality_report,
                requires_selection=False,
            )

        # Auto-fix medium issues (nulls, duplicates) silently
        if quality_report["summary"]["medium"] > 0:
            print(f"[quality] Auto-cleaning medium issues...")
            df = auto_clean_dataframe(df)
            print(f"[quality] auto-cleaned → {df.shape}")
            
            # Check if cleaning removed all rows
            if len(df) == 0:
                print(f"[quality] ERROR: All rows removed during cleaning!")
                raise ValueError("Data cleaning removed all rows. Please check your data quality.")

        # Generate session ID
        session_id = str(uuid.uuid4())

        # Store dataframe in session (file-based for production)
        save_session(session_id, file.filename, df)
        _cache.delete(session_id, "insight_result")
        _cache.delete(session_id, "insight_response")

        # Persist as a project in the database (legacy system — skip for new integer IDs)
        data_path = str(SESSION_DIR / session_id)
        if not session_id.isdigit():
            try:
                db.create_project(
                    project_id=session_id,
                    name=file.filename.rsplit('.', 1)[0],
                    filename=file.filename,
                    data_path=data_path,
                    row_count=len(df),
                    column_count=len(df.columns),
                )
                db.add_history(session_id, "upload", {"filename": file.filename, "rows": len(df), "cols": len(df.columns)}, f"Uploaded {file.filename}")
            except Exception as e:
                print(f"Warning: Could not persist project: {e}")
        
        # Build schema response
        columns_info = []
        for col in df.columns:
            col_data = df[col]
            null_count = col_data.null_count()
            columns_info.append(ColumnInfo(
                name=col,
                dtype=str(col_data.dtype),
                missing_count=null_count,
                missing_pct=round(null_count / len(df) * 100, 2) if len(df) > 0 else 0,
                unique_count=col_data.n_unique()
            ))
        
        # Generate preview (first 10 rows)
        preview = df.head(10).to_dicts()
        
        return DatasetSchema(
            session_id=session_id,
            filename=file.filename,
            row_count=len(df),
            column_count=len(df.columns),
            columns=columns_info,
            preview=preview,
            requires_selection=False,
            quality_report=quality_report,
        )
        
    except Exception as e:
        print(f"=== UPLOAD FAILED: {type(e).__name__}: {str(e)} ===")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=422, detail=f"Failed to process file: {str(e)}")

@app.get("/session/{session_id}")
def get_session(session_id: str):
    """Get info about an existing session."""
    try:
        filename, df = load_session(session_id)
    except FileNotFoundError:
        raise HTTPException(status_code=404, detail="Session not found")
    
    return {
        "session_id": session_id,
        "filename": filename,
        "row_count": len(df),
        "column_count": len(df.columns)
    }

@app.get("/data/{session_id}")
def get_raw_data(session_id: str, page: int = 1, page_size: int = 100):
    """Get paginated raw data for a session."""
    try:
        filename, df = load_session(session_id)
    except FileNotFoundError:
        raise HTTPException(status_code=404, detail="Session not found")
    
    # Limit page_size for performance
    page_size = min(page_size, 1000)
    page = max(1, page)
    
    total_rows = len(df)
    total_pages = (total_rows + page_size - 1) // page_size
    
    # Calculate slice indices
    start_idx = (page - 1) * page_size
    end_idx = min(start_idx + page_size, total_rows)
    
    # Get the slice of data
    data_slice = df.slice(start_idx, page_size).to_dicts()
    
    return {
        "session_id": session_id,
        "columns": df.columns,
        "data": data_slice,
        "page": page,
        "page_size": page_size,
        "total_rows": total_rows,
        "total_pages": total_pages
    }

# ============== PHASE 3: HEALTH CHECK & EDA ==============


class IssueItem(BaseModel):
    column: str
    issue_type: str  # "missing" | "duplicate" | "outlier" | "multivariate_outlier"
    severity: str    # "low" | "medium" | "high"
    description: str
    count: int
    percentage: float
    missingness_type: Optional[str] = None   # "MCAR" | "MAR" | "MNAR" | "TARGET_LEAKAGE"
    missingness_reason: Optional[str] = None
    columns_used: Optional[List[str]] = None  # populated for multivariate_outlier issues
    suggested_action: Optional[str] = None   # best CleanAction.action for this issue
    alternatives: Optional[List[str]] = None  # other valid actions
    group_col: Optional[str] = None          # set when group-wise imputation is recommended
    fill_col: Optional[str] = None           # set when fill-from-column is recommended

class HealthCheckResponse(BaseModel):
    session_id: str
    quality_score: str  # "A" | "B" | "C" | "D"
    row_count: int
    column_count: int
    duplicate_rows: int
    issues: list[IssueItem]

@app.get("/health-check/{session_id}", response_model=HealthCheckResponse)
def get_health_check(session_id: str):
    """
    Analyze dataset for data quality issues.
    This is the "Data Health Check" step - what a real Data Scientist does first.
    """
    try:
        filename, df = load_session(session_id)
    except FileNotFoundError:
        raise HTTPException(status_code=404, detail="Session not found")
    
    issues = []
    
    # 1. Check for duplicate rows
    duplicate_count = len(df) - df.unique().height
    if duplicate_count > 0:
        dup_pct = round(duplicate_count / len(df) * 100, 2)
        severity = "high" if dup_pct > 10 else ("medium" if dup_pct > 5 else "low")
        issues.append(IssueItem(
            column="_all_",
            issue_type="duplicate",
            severity=severity,
            description=f"{duplicate_count} duplicate rows detected ({dup_pct}%)",
            count=duplicate_count,
            percentage=dup_pct
        ))
    
    # 2. Check for missing values per column
    advisor = DataQualityAdvisor()
    for col in df.columns:
        null_count = df[col].null_count()
        if null_count > 0:
            null_pct = round(null_count / len(df) * 100, 2)
            severity = "high" if null_pct > 30 else ("medium" if null_pct > 10 else "low")
            mtype, mreason = advisor.classify_missingness(col, df)
            suggestion = advisor.suggest_action(col, df, mtype, null_pct)
            issues.append(IssueItem(
                column=col,
                issue_type="missing",
                severity=severity,
                description=f"{null_count} missing values ({null_pct}%)",
                count=null_count,
                percentage=null_pct,
                missingness_type=mtype,
                missingness_reason=mreason,
                suggested_action=suggestion["suggested_action"],
                alternatives=suggestion["alternatives"],
                group_col=suggestion["group_col"],
                fill_col=suggestion["fill_col"],
            ))
    
    # 3. Check for outliers in numeric columns (IQR method)
    numeric_cols = [col for col in df.columns if df[col].dtype in [pl.Float64, pl.Float32, pl.Int64, pl.Int32, pl.Int16, pl.Int8]]
    for col in numeric_cols:
        col_data = df[col].drop_nulls()
        if len(col_data) < 4:
            continue
        q1 = col_data.quantile(0.25)
        q3 = col_data.quantile(0.75)
        iqr = q3 - q1
        lower_bound = q1 - 1.5 * iqr
        upper_bound = q3 + 1.5 * iqr
        outlier_count = len(col_data.filter((col_data < lower_bound) | (col_data > upper_bound)))
        if outlier_count > 0:
            outlier_pct = round(outlier_count / len(col_data) * 100, 2)
            severity = "high" if outlier_pct > 10 else ("medium" if outlier_pct > 5 else "low")
            issues.append(IssueItem(
                column=col,
                issue_type="outlier",
                severity=severity,
                description=f"{outlier_count} outliers detected ({outlier_pct}%)",
                count=outlier_count,
                percentage=outlier_pct
            ))
    
    # 4. Multivariate outlier detection (IsolationForest — Gap 4)
    for mv in advisor.multivariate_outlier_check(df):
        issues.append(IssueItem(
            column="_multivariate_",
            issue_type=mv["issue_type"],
            severity=mv["severity"],
            description=mv["reason"],
            count=mv["count"],
            percentage=round(mv["count"] / len(df) * 100, 2),
            columns_used=mv["columns_used"],
        ))

    # Calculate quality score
    high_issues = sum(1 for i in issues if i.severity == "high")
    medium_issues = sum(1 for i in issues if i.severity == "medium")
    
    if high_issues >= 3:
        quality_score = "D"
    elif high_issues >= 1:
        quality_score = "C"
    elif medium_issues >= 3:
        quality_score = "B"
    else:
        quality_score = "A"
    
    return HealthCheckResponse(
        session_id=session_id,
        quality_score=quality_score,
        row_count=len(df),
        column_count=len(df.columns),
        duplicate_rows=duplicate_count,
        issues=issues
    )

class CleanAction(BaseModel):
    action: str  # "drop_duplicates" | "drop_column" | "impute_mean" | "impute_median" | "impute_mode" | "drop_nulls" | "impute_group_median" | "impute_from_col"
    column: Optional[str] = None
    group_col: Optional[str] = None   # for impute_group_median (Gap 6)
    fill_col: Optional[str] = None    # for impute_from_col (Gap 7)
    enabled: bool = True  # Allow toggling actions

class CleaningPreview(BaseModel):
    before_rows: int
    before_columns: int
    before_score: str
    after_rows: int
    after_columns: int
    after_score: str
    changes: list[str]
    sample_affected: list[dict]

def calculate_quality_score(df: pl.DataFrame) -> str:
    """Calculate quality score for a DataFrame."""
    total_cells = len(df) * len(df.columns)
    if total_cells == 0:
        return "A"
    
    null_count = sum(df[col].null_count() for col in df.columns)
    null_pct = (null_count / total_cells) * 100
    
    duplicate_count = len(df) - len(df.unique())
    dup_pct = (duplicate_count / len(df)) * 100 if len(df) > 0 else 0
    
    if null_pct > 20 or dup_pct > 10:
        return "D"
    elif null_pct > 10 or dup_pct > 5:
        return "C"
    elif null_pct > 5 or dup_pct > 2:
        return "B"
    return "A"

def impute_with_flag(df: pl.DataFrame, col: str, fill_value) -> pl.DataFrame:
    """Impute nulls in *col* and add a companion {col}_was_missing Int8 flag.

    The flag is created BEFORE filling so it captures the original null
    pattern — preserving MNAR signal for any downstream model.
    If fill_value is None (e.g. all-null column), the flag is still created
    but the fill is skipped.
    """
    flag_col = f"{col}_was_missing"
    df = df.with_columns(pl.col(col).is_null().cast(pl.Int8).alias(flag_col))
    if fill_value is not None:
        df = df.with_columns(pl.col(col).fill_null(fill_value))
    return df


class DataQualityAdvisor:
    """Diagnose *why* values are missing (MCAR / MAR / MNAR).

    Gap 2 implementation: semantic keyword heuristic first, then
    Pearson correlation between the is_missing indicator and every
    other numeric column.
    """

    _MNAR_KEYWORDS = [
        "delivery", "delivered", "shipped", "dispatch",
        "return", "returned", "refund", "resolved",
        "closed", "completed", "end_date", "discharge",
        "death", "churn", "churned", "cancel", "cancelled",
    ]

    def classify_missingness(
        self,
        col: str,
        df: pl.DataFrame,
        target: Optional[str] = None,
    ) -> tuple[str, str]:
        """Return (mechanism, reason_string) for *col*.

        mechanism is one of 'MCAR', 'MAR', 'MNAR', 'TARGET_LEAKAGE'.
        TARGET_LEAKAGE overrides all other mechanisms when the missingness
        indicator correlates with the target column (|r| > 0.3).
        """
        col_lower = col.lower()

        # ── 0. Target-conditional check (Gap 13) ────────────────────────
        if target and target != col and target in df.columns:
            try:
                pdf = df.to_pandas()
                missing_flag = pdf[col].isna().astype(int)
                if missing_flag.sum() > 0:
                    target_col = pdf[target]
                    max_r = 0.0
                    if pd.api.types.is_numeric_dtype(target_col):
                        r = missing_flag.corr(target_col.fillna(target_col.median()))
                        if pd.notna(r):
                            max_r = abs(float(r))
                    else:
                        n_unique = target_col.nunique()
                        if 2 <= n_unique <= 20:
                            dummies = pd.get_dummies(target_col.fillna("_missing_"))
                            for dc in dummies.columns:
                                r = missing_flag.corr(dummies[dc].astype(float))
                                if pd.notna(r) and abs(float(r)) > max_r:
                                    max_r = abs(float(r))
                    if max_r > 0.3:
                        return (
                            "TARGET_LEAKAGE",
                            f"Missingness predicts target '{target}' (|r|={max_r:.2f}). "
                            "Imputing this column will inject a label-leaking signal into the model.",
                        )
            except Exception:
                pass

        # ── 1. Semantic MNAR ────────────────────────────────────────────
        if any(k in col_lower for k in self._MNAR_KEYWORDS):
            return (
                "MNAR",
                f"Missing likely means the event '{col}' has not yet occurred "
                f"(Missing Not At Random — do not impute).",
            )

        # ── 2. Correlation-based MAR ────────────────────────────────────
        try:
            pdf = df.to_pandas()
            missing_flag = pdf[col].isna().astype(int)
            if missing_flag.sum() == 0:
                return "MCAR", "Column has no missing values."

            num_cols = [
                c for c in pdf.columns
                if c != col and pd.api.types.is_numeric_dtype(pdf[c])
            ][:20]  # cap at 20 to keep it fast

            correlates: dict[str, float] = {}
            for other in num_cols:
                try:
                    r = missing_flag.corr(pdf[other].fillna(0))
                    if pd.notna(r) and abs(r) > 0.3:
                        correlates[other] = round(float(r), 2)
                except Exception:
                    pass

            if correlates:
                top = sorted(correlates.items(), key=lambda x: abs(x[1]), reverse=True)[:3]
                detail = ", ".join(f"{k}={v:+.2f}" for k, v in top)
                return "MAR", f"Missingness correlates with: {detail}"
        except Exception:
            pass

        # ── 3. Default MCAR ─────────────────────────────────────────────
        return "MCAR", "No strong correlation with other features detected."

    def suggest_action(
        self,
        col: str,
        df: pl.DataFrame,
        missingness_type: str,
        null_pct: float,
    ) -> dict:
        """Return the best CleanAction fields for a missing-values IssueItem.

        Considers missingness mechanism, column dtype, related columns, and the
        discount/price keyword heuristic from Gap 7.
        """
        _NUMERIC = (pl.Float64, pl.Float32, pl.Int64, pl.Int32, pl.Int16, pl.Int8)
        is_numeric = df[col].dtype in _NUMERIC

        # ── detect a useful group_col (low-cardinality cat correlated with missingness) ──
        group_col_hint: Optional[str] = None
        try:
            pdf = df.to_pandas()
            missing_flag = pdf[col].isna().astype(int)
            if missing_flag.sum() > 0:
                best_r = 0.0
                for c in df.columns:
                    if c == col:
                        continue
                    if df[c].dtype not in (pl.Utf8, pl.Boolean, pl.Categorical):
                        continue
                    n_uniq = df[c].n_unique()
                    if not (2 <= n_uniq <= 20):
                        continue
                    try:
                        dummies = pd.get_dummies(pdf[c].fillna("_missing_"))
                        for dc in dummies.columns:
                            r = missing_flag.corr(dummies[dc].astype(float))
                            if pd.notna(r) and abs(float(r)) > best_r:
                                best_r = abs(float(r))
                                group_col_hint = c
                    except Exception:
                        pass
                if best_r < 0.2:
                    group_col_hint = None
        except Exception:
            pass

        # ── detect fill_col (discount/original price keyword pair) ──
        fill_col_hint: Optional[str] = None
        col_lower = col.lower()
        if any(k in col_lower for k in _DISCOUNT_COL_KWS):
            for other in df.columns:
                if other != col and any(k in other.lower() for k in _ORIGINAL_COL_KWS):
                    fill_col_hint = other
                    break

        # ── decision tree ────────────────────────────────────────────────
        alts_numeric = ["impute_mean", "impute_mode", "drop_nulls"]
        alts_cat = ["drop_nulls", "drop_column"]

        if missingness_type == "TARGET_LEAKAGE":
            return dict(
                suggested_action="drop_column",
                alternatives=["impute_median"] if is_numeric else ["impute_mode"],
                group_col=None, fill_col=None,
            )

        if missingness_type == "MNAR":
            action = "drop_column" if null_pct > 40 else "drop_nulls"
            alts = ["impute_mode", "impute_median"] if is_numeric else ["impute_mode"]
            return dict(suggested_action=action, alternatives=alts,
                        group_col=None, fill_col=None)

        if null_pct > 70:
            return dict(
                suggested_action="drop_column",
                alternatives=["impute_median"] if is_numeric else ["impute_mode"],
                group_col=None, fill_col=fill_col_hint,
            )

        if fill_col_hint:
            return dict(
                suggested_action="impute_from_col",
                alternatives=["impute_median", "impute_mode"],
                group_col=None, fill_col=fill_col_hint,
            )

        if group_col_hint and missingness_type == "MAR":
            action = "impute_group_median" if is_numeric else "impute_mode"
            return dict(
                suggested_action=action,
                alternatives=["impute_median", "impute_mode"],
                group_col=group_col_hint, fill_col=None,
            )

        if is_numeric:
            return dict(
                suggested_action="impute_median",
                alternatives=alts_numeric,
                group_col=group_col_hint, fill_col=None,
            )

        return dict(
            suggested_action="impute_mode",
            alternatives=alts_cat,
            group_col=None, fill_col=None,
        )

    def multivariate_outlier_check(self, df: pl.DataFrame) -> list[dict]:
        """Flag rows anomalous in combination using IsolationForest (Gap 4).

        Only runs when sklearn is available and dataset has ≥50 complete rows
        across at least two numeric columns.
        """
        try:
            from sklearn.ensemble import IsolationForest
        except ImportError:
            return []

        numeric_cols = [
            c for c in df.columns
            if df[c].dtype in (pl.Float64, pl.Float32, pl.Int64, pl.Int32, pl.Int16, pl.Int8)
        ][:10]

        if len(numeric_cols) < 2:
            return []

        pdf = df[numeric_cols].to_pandas().dropna()
        if len(pdf) < 50:
            return []

        clf = IsolationForest(contamination=0.05, random_state=42)
        preds = clf.fit_predict(pdf)
        anomaly_count = int((preds == -1).sum())

        if anomaly_count == 0:
            return []

        return [{
            "issue_type": "multivariate_outlier",
            "count": anomaly_count,
            "severity": "medium",
            "reason": (
                "IsolationForest detected rows with unusual value combinations. "
                "Flag for review — do not auto-remove."
            ),
            "columns_used": numeric_cols,
        }]


class SchemaSnapshot:
    """Capture training-time schema for drift detection at inference time (Gap 9).

    Both methods are static — persist the dict from from_df() inside a CleaningRecipe
    and pass it back to validate() against any future DataFrame.
    """

    _MAX_CAT_UNIQUE = 50  # columns with ≤ this many uniques are treated as categorical

    @staticmethod
    def from_df(df: pl.DataFrame) -> dict:
        """Build a serializable schema snapshot from a training DataFrame."""
        _NUMERIC = (pl.Float64, pl.Float32, pl.Int64, pl.Int32, pl.Int16, pl.Int8)

        columns = list(df.columns)
        dtypes = {c: str(df[c].dtype) for c in columns}

        categorical_levels: dict[str, list] = {}
        for c in columns:
            if df[c].dtype in (pl.Utf8, pl.Boolean, pl.Categorical):
                if df[c].n_unique() <= SchemaSnapshot._MAX_CAT_UNIQUE:
                    categorical_levels[c] = sorted(
                        str(v) for v in df[c].drop_nulls().unique().to_list()
                    )

        numeric_ranges: dict[str, dict] = {}
        numeric_quantiles: dict[str, list] = {}
        _PSI_PERCENTILES = (0.01, 0.05, 0.10, 0.25, 0.50, 0.75, 0.90, 0.95, 0.99)

        for c in columns:
            if df[c].dtype in _NUMERIC:
                col_data = df[c].drop_nulls()
                if len(col_data) > 0:
                    numeric_ranges[c] = {
                        "min": float(col_data.min()),
                        "max": float(col_data.max()),
                    }
                if len(col_data) >= 10:
                    numeric_quantiles[c] = [
                        float(col_data.quantile(q)) for q in _PSI_PERCENTILES
                    ]

        return {
            "columns": columns,
            "dtypes": dtypes,
            "categorical_levels": categorical_levels,
            "numeric_ranges": numeric_ranges,
            "numeric_quantiles": numeric_quantiles,
            "train_rows": len(df),
        }

    # Training bucket proportions matching [P1,P5,P10,P25,P50,P75,P90,P95,P99]
    _PSI_TRAIN_PCTS = [0.01, 0.04, 0.05, 0.15, 0.25, 0.25, 0.15, 0.05, 0.04, 0.01]

    @staticmethod
    def _compute_psi(actual_values: pl.Series, training_quantiles: list[float]) -> float:
        """Population Stability Index using training quantile bins (Gap 12).

        Bins: (-inf,P1), [P1,P5), [P5,P10), [P10,P25), [P25,P50),
              [P50,P75), [P75,P90), [P90,P95), [P95,P99), [P99,+inf)

        Industry thresholds: <0.1 stable, 0.1-0.25 moderate, >=0.25 significant.
        """
        import math
        eps = 1e-4
        vals = actual_values.drop_nulls().to_list()
        n = len(vals)
        if n == 0:
            return 0.0

        boundaries = [float("-inf")] + list(training_quantiles) + [float("inf")]
        counts = [0] * 10
        for v in vals:
            for i in range(10):
                if boundaries[i] <= v < boundaries[i + 1]:
                    counts[i] += 1
                    break

        psi = 0.0
        for i in range(10):
            actual_pct = max(counts[i] / n, eps)
            train_pct  = max(SchemaSnapshot._PSI_TRAIN_PCTS[i], eps)
            psi += (actual_pct - train_pct) * math.log(actual_pct / train_pct)
        return psi

    @staticmethod
    def validate(df: pl.DataFrame, snapshot: dict) -> list[dict]:
        """Return a list of schema violation dicts comparing df against the snapshot.

        Checks:
        1. Missing columns (high)
        2. Dtype changes (high)
        3. Unseen categorical levels (medium)
        4. Hard range breach (medium) — kept as complement to PSI
        5. PSI-based numeric drift (medium / high) — replaces the 2-std mean-shift check
        """
        _NUMERIC = (pl.Float64, pl.Float32, pl.Int64, pl.Int32, pl.Int16, pl.Int8)
        violations: list[dict] = []

        snap_cols: list[str] = snapshot.get("columns", [])
        snap_dtypes: dict = snapshot.get("dtypes", {})
        snap_levels: dict = snapshot.get("categorical_levels", {})
        snap_ranges: dict = snapshot.get("numeric_ranges", {})
        snap_quantiles: dict = snapshot.get("numeric_quantiles", {})

        # 1. Missing columns
        for col in snap_cols:
            if col not in df.columns:
                violations.append({
                    "column": col,
                    "issue": "missing_column",
                    "severity": "high",
                    "detail": f"Column '{col}' was present at training but is absent.",
                })

        # 2. Dtype changes
        for col, expected in snap_dtypes.items():
            if col in df.columns:
                actual = str(df[col].dtype)
                if actual != expected:
                    violations.append({
                        "column": col,
                        "issue": "dtype_changed",
                        "severity": "high",
                        "expected": expected,
                        "actual": actual,
                        "detail": f"Column '{col}' dtype changed from {expected} to {actual}.",
                    })

        # 3. Unseen categorical levels
        for col, known in snap_levels.items():
            if col in df.columns:
                known_set = set(known)
                new_vals = sorted(
                    str(v) for v in df[col].drop_nulls().unique().to_list()
                    if str(v) not in known_set
                )
                if new_vals:
                    violations.append({
                        "column": col,
                        "issue": "new_categories",
                        "severity": "medium",
                        "new_values": new_vals[:20],
                        "detail": f"{len(new_vals)} unseen category value(s) in '{col}'.",
                    })

        # 4. Hard range breach (kept — complementary to PSI, catches boundary violations)
        for col, rng in snap_ranges.items():
            if col not in df.columns or df[col].dtype not in _NUMERIC:
                continue
            col_data = df[col].drop_nulls()
            if len(col_data) == 0:
                continue
            actual_min = float(col_data.min())
            actual_max = float(col_data.max())
            if actual_min < rng["min"] or actual_max > rng["max"]:
                violations.append({
                    "column": col,
                    "issue": "range_breach",
                    "severity": "medium",
                    "training_range": [rng["min"], rng["max"]],
                    "actual_range": [actual_min, actual_max],
                    "detail": f"Values in '{col}' fall outside the training range.",
                })

        # 5. PSI-based drift (replaces 2-std mean-shift — catches bimodal/tail shifts)
        for col, quantiles in snap_quantiles.items():
            if col not in df.columns or df[col].dtype not in _NUMERIC:
                continue
            col_data = df[col].drop_nulls()
            if len(col_data) == 0:
                continue
            psi = SchemaSnapshot._compute_psi(col_data, quantiles)
            if psi >= 0.25:
                violations.append({
                    "column": col,
                    "issue": "significant_drift",
                    "severity": "high",
                    "psi_value": round(psi, 4),
                    "training_quantiles": quantiles,
                    "detail": (
                        f"PSI={psi:.4f} ≥ 0.25 — significant distribution shift in '{col}'. "
                        "Model scores are likely unreliable; retraining is recommended."
                    ),
                })
            elif psi >= 0.1:
                violations.append({
                    "column": col,
                    "issue": "moderate_drift",
                    "severity": "medium",
                    "psi_value": round(psi, 4),
                    "training_quantiles": quantiles,
                    "detail": (
                        f"PSI={psi:.4f} in [0.1, 0.25) — moderate distribution shift in '{col}'. "
                        "Monitor closely; consider retraining."
                    ),
                })

        return violations


class CleaningRecipe:
    """Fit a cleaning recipe on training data; apply it to any split without refitting (Gap 3).

    Usage pattern:
        recipe = CleaningRecipe.fit(train_df, actions)   # compute stats from train only
        train_clean, _ = CleaningRecipe.apply(train_df, recipe)
        test_clean,  _ = CleaningRecipe.apply(test_df,  recipe)  # same values, no leakage
    """

    @staticmethod
    def fit(df: pl.DataFrame, actions: list["CleanAction"]) -> dict:
        """Compute all imputation statistics from df and return a serializable recipe dict."""
        from datetime import datetime, timezone

        steps: list[dict] = []
        for action in actions:
            if not action.enabled:
                continue
            step: dict = {"action": action.action, "from_training": True}
            if action.column:
                step["column"] = action.column

            col = action.column
            if action.action == "impute_mean" and col and col in df.columns:
                step["fill_value"] = df[col].mean()
            elif action.action == "impute_median" and col and col in df.columns:
                step["fill_value"] = df[col].median()
            elif action.action == "impute_mode" and col and col in df.columns:
                mode_list = df[col].mode().to_list()
                step["fill_value"] = mode_list[0] if mode_list else None
            elif (
                action.action == "impute_group_median"
                and col and col in df.columns
                and action.group_col and action.group_col in df.columns
            ):
                step["group_col"] = action.group_col
                step["global_median"] = df[col].median()
                gm_df = df.group_by(action.group_col).agg(
                    pl.col(col).median().alias("_gm")
                )
                # Serialize as {str(group_val): median} — JSON keys must be strings
                step["group_medians"] = {
                    str(row[action.group_col]): row["_gm"]
                    for row in gm_df.to_dicts()
                }
            elif action.action == "impute_from_col" and col and action.fill_col:
                step["fill_col"] = action.fill_col
                # Pre-compute discount-flag decision so apply() is deterministic
                step["creates_had_discount"] = (
                    any(k in col.lower() for k in _DISCOUNT_COL_KWS)
                    and any(k in action.fill_col.lower() for k in _ORIGINAL_COL_KWS)
                )

            steps.append(step)

        return {
            "version": 1,
            "fitted_at": datetime.now(timezone.utc).isoformat(),
            "train_rows": len(df),
            "steps": steps,
            "schema_snapshot": SchemaSnapshot.from_df(df),  # Gap 9
        }

    @staticmethod
    def apply(df: pl.DataFrame, recipe: dict) -> tuple[pl.DataFrame, list[str]]:
        """Apply a pre-fitted recipe to df. Never recomputes imputation statistics."""
        changelog: list[str] = []

        for step in recipe.get("steps", []):
            action = step.get("action")
            col = step.get("column")
            fill_value = step.get("fill_value")

            if action == "drop_duplicates":
                before = len(df)
                df = df.unique()
                removed = before - len(df)
                if removed > 0:
                    changelog.append(f"Removed {removed} duplicate rows")

            elif action == "drop_column" and col:
                if col in df.columns:
                    df = df.drop(col)
                    changelog.append(f"Dropped column '{col}'")

            elif action == "drop_nulls" and col:
                before = len(df)
                df = df.filter(pl.col(col).is_not_null())
                removed = before - len(df)
                if removed > 0:
                    changelog.append(f"Dropped {removed} rows with null '{col}'")

            elif action in ("impute_mean", "impute_median", "impute_mode") and col and fill_value is not None:
                if col in df.columns:
                    null_count = df[col].null_count()
                    if null_count > 0:
                        df = impute_with_flag(df, col, fill_value)
                        flag_col = f"{col}_was_missing"
                        stat_name = action.replace("impute_", "")
                        changelog.append(
                            f"Imputed {null_count} missing '{col}' with {stat_name} "
                            f"({fill_value:.4g} from training); added '{flag_col}' flag"
                        )

            elif action == "impute_group_median" and col:
                group_col = step.get("group_col")
                group_medians_map: dict = step.get("group_medians", {})
                global_median = step.get("global_median")
                if col in df.columns and group_col and group_col in df.columns:
                    null_count = df[col].null_count()
                    if null_count > 0:
                        flag_col = f"{col}_was_missing"
                        df = df.with_columns(pl.col(col).is_null().cast(pl.Int8).alias(flag_col))
                        # Reconstruct lookup DataFrame from serialized training medians
                        gm_df = pl.DataFrame({
                            group_col: list(group_medians_map.keys()),
                            "_group_median": [
                                float(v) if v is not None else None
                                for v in group_medians_map.values()
                            ],
                        })
                        try:
                            gm_df = gm_df.with_columns(
                                pl.col(group_col).cast(df[group_col].dtype)
                            )
                        except Exception:
                            pass
                        df = df.join(gm_df, on=group_col, how="left")
                        df = df.with_columns(
                            pl.when(pl.col(col).is_null())
                            .then(pl.col("_group_median").fill_null(global_median))
                            .otherwise(pl.col(col))
                            .alias(col)
                        ).drop("_group_median")
                        changelog.append(
                            f"Imputed {null_count} missing '{col}' with group-wise median "
                            f"(by '{group_col}', fitted on training); added '{flag_col}' flag"
                        )

            elif action == "impute_from_col" and col:
                fill_col = step.get("fill_col")
                creates_had_discount = step.get("creates_had_discount", False)
                if col in df.columns and fill_col and fill_col in df.columns:
                    null_count = df[col].null_count()
                    if null_count > 0:
                        flag_col = f"{col}_was_missing"
                        df = df.with_columns(pl.col(col).is_null().cast(pl.Int8).alias(flag_col))
                        if creates_had_discount:
                            df = df.with_columns(
                                pl.when(pl.col(col).is_null())
                                .then(pl.lit(0, dtype=pl.Int8))
                                .otherwise(pl.lit(1, dtype=pl.Int8))
                                .alias("had_discount")
                            )
                        df = df.with_columns(
                            pl.when(pl.col(col).is_null())
                            .then(pl.col(fill_col))
                            .otherwise(pl.col(col))
                            .alias(col)
                        )
                        msg = (
                            f"Imputed {null_count} missing '{col}' from '{fill_col}' "
                            f"(recipe, no refit); added '{flag_col}' flag"
                        )
                        if creates_had_discount:
                            msg += (
                                "; added 'had_discount' transparency flag "
                                "(1 = item had a discounted price, 0 = no discount was applied)"
                            )
                        changelog.append(msg)

        return df, changelog


# ── Gap 11: Pluggable split strategies ──────────────────────────────────────

def split_dataframe(
    df: pl.DataFrame,
    strategy: str,
    train_fraction: float,
    random_state: int = 42,
    time_col: Optional[str] = None,
    stratify_col: Optional[str] = None,
    group_col: Optional[str] = None,
) -> tuple[pl.DataFrame, pl.DataFrame]:
    """Return (train_df, test_df) using the requested strategy.

    Raises HTTPException(422) when a required column is missing or the
    strategy name is unknown.
    """
    _NUMERIC_DTYPES = (pl.Float64, pl.Float32, pl.Int64, pl.Int32, pl.Int16, pl.Int8)

    if strategy == "positional":
        idx = int(len(df) * train_fraction)
        return df[:idx], df[idx:]

    elif strategy == "random":
        shuffled = df.sample(fraction=1.0, shuffle=True, seed=random_state)
        idx = int(len(shuffled) * train_fraction)
        return shuffled[:idx], shuffled[idx:]

    elif strategy == "time":
        if not time_col or time_col not in df.columns:
            raise HTTPException(
                status_code=422,
                detail=f"strategy='time' requires time_col; "
                       f"'{time_col}' not found in dataset.",
            )
        sorted_df = df.sort(time_col, nulls_last=True)
        idx = int(len(sorted_df) * train_fraction)
        return sorted_df[:idx], sorted_df[idx:]

    elif strategy == "stratified":
        if not stratify_col or stratify_col not in df.columns:
            raise HTTPException(
                status_code=422,
                detail=f"strategy='stratified' requires stratify_col; "
                       f"'{stratify_col}' not found in dataset.",
            )
        strat_col = stratify_col
        df_work = df
        # Bin high-cardinality numeric columns into quartiles
        if df[strat_col].dtype in _NUMERIC_DTYPES and df[strat_col].n_unique() > 20:
            pdf = df.to_pandas()
            pdf["_strat_bin"] = pd.qcut(
                pdf[strat_col].fillna(pdf[strat_col].median()),
                q=4, labels=False, duplicates="drop",
            )
            df_work = pl.from_pandas(pdf)
            strat_col = "_strat_bin"

        train_parts: list[pl.DataFrame] = []
        test_parts: list[pl.DataFrame] = []
        for gval in df_work[strat_col].unique().to_list():
            if gval is None:
                continue
            stratum = df_work.filter(pl.col(strat_col) == gval)
            idx = max(1, int(len(stratum) * train_fraction))
            train_parts.append(stratum[:idx])
            if len(stratum) > idx:
                test_parts.append(stratum[idx:])

        train_df = pl.concat(train_parts) if train_parts else df_work[:0]
        test_df  = pl.concat(test_parts)  if test_parts  else df_work[:0]

        # Drop temporary bin column if added
        for col in ("_strat_bin",):
            if col in train_df.columns:
                train_df = train_df.drop(col)
            if col in test_df.columns:
                test_df = test_df.drop(col)
        return train_df, test_df

    elif strategy == "group":
        if not group_col or group_col not in df.columns:
            raise HTTPException(
                status_code=422,
                detail=f"strategy='group' requires group_col; "
                       f"'{group_col}' not found in dataset.",
            )
        import random as _random
        rng = _random.Random(random_state)
        all_groups = df[group_col].drop_nulls().unique().to_list()
        rng.shuffle(all_groups)
        n_train = max(1, int(len(all_groups) * train_fraction))
        train_groups = set(all_groups[:n_train])
        test_groups  = set(all_groups[n_train:])
        return (
            df.filter(pl.col(group_col).is_in(list(train_groups))),
            df.filter(pl.col(group_col).is_in(list(test_groups))),
        )

    else:
        raise HTTPException(status_code=422, detail=f"Unknown split_strategy '{strategy}'.")


class FitRecipeRequest(BaseModel):
    actions: list[CleanAction]
    train_fraction: float = 0.8
    split_strategy: Literal["positional", "random", "time", "stratified", "group"] = "positional"
    random_state: int = 42
    time_col: Optional[str] = None       # required when strategy="time"
    stratify_col: Optional[str] = None   # required when strategy="stratified"
    group_col: Optional[str] = None      # required when strategy="group"


class ApplyRecipeRequest(BaseModel):
    recipe: dict


def _group_wise_median(df: pl.DataFrame, col: str, group_col: str) -> pl.DataFrame:
    """Impute col in-place using per-group medians; global median fallback for unseen groups."""
    global_median = df[col].median()
    group_medians = (
        df.group_by(group_col)
        .agg(pl.col(col).median().alias("_group_median"))
    )
    df = df.join(group_medians, on=group_col, how="left")
    df = df.with_columns(
        pl.when(pl.col(col).is_null())
        .then(pl.col("_group_median").fill_null(global_median))
        .otherwise(pl.col(col))
        .alias(col)
    )
    return df.drop("_group_median")


def apply_actions_to_df(df: pl.DataFrame, actions: list[CleanAction]) -> tuple[pl.DataFrame, list[str]]:
    """Apply cleaning actions and return cleaned df with changelog."""
    changelog = []
    
    for action in actions:
        if not action.enabled:
            continue
            
        if action.action == "drop_duplicates":
            before = len(df)
            df = df.unique()
            removed = before - len(df)
            if removed > 0:
                changelog.append(f"Removed {removed} duplicate rows")
        elif action.action == "drop_column" and action.column:
            if action.column in df.columns:
                df = df.drop(action.column)
                changelog.append(f"Dropped column '{action.column}'")
        elif action.action == "drop_nulls" and action.column:
            before = len(df)
            df = df.filter(pl.col(action.column).is_not_null())
            removed = before - len(df)
            if removed > 0:
                changelog.append(f"Dropped {removed} rows with null '{action.column}'")
        elif action.action == "impute_mean" and action.column:
            col = action.column
            if col not in df.columns:
                continue
            null_count = df[col].null_count()
            if null_count > 0:
                mean_val = df[col].mean()
                df = impute_with_flag(df, col, mean_val)
                flag_col = f"{col}_was_missing"
                if mean_val is not None:
                    changelog.append(
                        f"Imputed {null_count} missing '{col}' with mean ({mean_val:.2f}); "
                        f"added '{flag_col}' flag"
                    )
                else:
                    changelog.append(
                        f"Flagged {null_count} missing '{col}' (all values null — no fill); "
                        f"added '{flag_col}'"
                    )
        elif action.action == "impute_median" and action.column:
            col = action.column
            if col not in df.columns:
                continue
            null_count = df[col].null_count()
            if null_count > 0:
                median_val = df[col].median()
                df = impute_with_flag(df, col, median_val)
                flag_col = f"{col}_was_missing"
                if median_val is not None:
                    changelog.append(
                        f"Imputed {null_count} missing '{col}' with median ({median_val:.2f}); "
                        f"added '{flag_col}' flag"
                    )
                else:
                    changelog.append(
                        f"Flagged {null_count} missing '{col}' (all values null — no fill); "
                        f"added '{flag_col}'"
                    )
        elif action.action == "impute_mode" and action.column:
            col = action.column
            if col not in df.columns:
                continue
            null_count = df[col].null_count()
            if null_count > 0:
                mode_list = df[col].mode().to_list()
                mode_val = mode_list[0] if len(mode_list) > 0 else None
                if mode_val is not None:
                    df = impute_with_flag(df, col, mode_val)
                    flag_col = f"{col}_was_missing"
                    changelog.append(
                        f"Imputed {null_count} missing '{col}' with mode ('{mode_val}'); "
                        f"added '{flag_col}' flag"
                    )
        elif action.action == "impute_group_median" and action.column and action.group_col:
            col = action.column
            grp = action.group_col
            if col in df.columns and grp in df.columns:
                null_count = df[col].null_count()
                if null_count > 0:
                    flag_col = f"{col}_was_missing"
                    df = df.with_columns(pl.col(col).is_null().cast(pl.Int8).alias(flag_col))
                    df = _group_wise_median(df, col, grp)
                    changelog.append(
                        f"Imputed {null_count} missing '{col}' with group-wise median "
                        f"(grouped by '{grp}'); added '{flag_col}' flag"
                    )
        elif action.action == "impute_from_col" and action.column and action.fill_col:
            col = action.column
            fill_col = action.fill_col
            if col in df.columns and fill_col in df.columns:
                null_count = df[col].null_count()
                if null_count > 0:
                    flag_col = f"{col}_was_missing"
                    df = df.with_columns(pl.col(col).is_null().cast(pl.Int8).alias(flag_col))
                    # Gap 7: discount transparency flag — created BEFORE filling
                    creates_had_discount = (
                        any(k in col.lower() for k in _DISCOUNT_COL_KWS)
                        and any(k in fill_col.lower() for k in _ORIGINAL_COL_KWS)
                    )
                    if creates_had_discount:
                        df = df.with_columns(
                            pl.when(pl.col(col).is_null())
                            .then(pl.lit(0, dtype=pl.Int8))
                            .otherwise(pl.lit(1, dtype=pl.Int8))
                            .alias("had_discount")
                        )
                    df = df.with_columns(
                        pl.when(pl.col(col).is_null())
                        .then(pl.col(fill_col))
                        .otherwise(pl.col(col))
                        .alias(col)
                    )
                    msg = (
                        f"Imputed {null_count} missing '{col}' from '{fill_col}'; "
                        f"added '{flag_col}' flag"
                    )
                    if creates_had_discount:
                        msg += (
                            "; added 'had_discount' transparency flag — "
                            "A 'had_discount' column was created: "
                            "models can use this signal explicitly "
                            "(1 = item had a discounted price, 0 = no discount was applied)"
                        )
                    changelog.append(msg)

    return df, changelog

@app.post("/preview-clean/{session_id}")
def preview_cleaning(session_id: str, actions: list[CleanAction]):
    """
    Preview cleaning actions without applying them.
    Returns before/after stats and sample of affected rows.
    """
    try:
        filename, df = load_session(session_id)
    except FileNotFoundError:
        raise HTTPException(status_code=404, detail="Session not found")
    
    # Before stats
    before_rows = len(df)
    before_columns = len(df.columns)
    before_score = calculate_quality_score(df)
    
    # Apply to a copy
    df_copy = df.clone()
    cleaned_df, changelog = apply_actions_to_df(df_copy, actions)
    
    # After stats
    after_rows = len(cleaned_df)
    after_columns = len(cleaned_df.columns)
    after_score = calculate_quality_score(cleaned_df)
    
    # Get sample of affected rows (first 5)
    sample_affected = []
    try:
        sample_affected = cleaned_df.head(5).to_dicts()
    except Exception:
        pass
    
    return {
        "before_rows": before_rows,
        "before_columns": before_columns,
        "before_score": before_score,
        "after_rows": after_rows,
        "after_columns": after_columns, 
        "after_score": after_score,
        "row_delta": after_rows - before_rows,
        "column_delta": after_columns - before_columns,
        "changes": changelog,
        "sample_affected": sample_affected
    }

@app.post("/clean/{session_id}")
def apply_cleaning(session_id: str, actions: list[CleanAction]):
    """
    Apply cleaning actions to the dataset.
    Saves backup of original data for undo capability.
    """
    try:
        filename, df = load_session(session_id)
    except FileNotFoundError:
        raise HTTPException(status_code=404, detail="Session not found")
    
    # Save backup before cleaning
    backup_path = SESSION_DIR / session_id / "backup.parquet"
    try:
        df.write_parquet(backup_path)
    except Exception as e:
        print(f"Warning: Could not save backup: {e}")
    
    # Apply cleaning actions
    cleaned_df, changelog = apply_actions_to_df(df, actions)
    
    # Save the cleaned DataFrame back to disk
    save_session(session_id, filename, cleaned_df)
    
    # Track cleaning history
    if not session_id.isdigit():
        try:
            db.add_history(
                session_id, "clean",
                {"actions": [a.dict() for a in actions]},
                f"Applied {len(actions)} cleaning actions: {', '.join(changelog[:3])}"
            )
            db.update_project(session_id, row_count=len(cleaned_df), column_count=len(cleaned_df.columns))
        except Exception:
            pass
    
    return {
        "status": "ok", 
        "row_count": len(cleaned_df), 
        "column_count": len(cleaned_df.columns),
        "changes": changelog,
        "can_undo": backup_path.exists()
    }

@app.post("/undo-clean/{session_id}")
def undo_cleaning(session_id: str):
    """
    Restore the original data from backup.
    """
    try:
        filename, _ = load_session(session_id)
    except FileNotFoundError:
        raise HTTPException(status_code=404, detail="Session not found")

    backup_path = SESSION_DIR / session_id / "backup.parquet"

    if not backup_path.exists():
        raise HTTPException(status_code=404, detail="No backup available to restore")

    try:
        # Restore from backup
        df = pl.read_parquet(backup_path)
        save_session(session_id, filename, df)

        # Remove backup after restore
        backup_path.unlink()

        return {
            "status": "ok",
            "message": "Data restored to original state",
            "row_count": len(df),
            "column_count": len(df.columns)
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to restore backup: {str(e)}")


# ── Gap 3: Train/test-aware cleaning endpoints ──────────────────────────────

@app.post("/clean/fit/{session_id}")
def fit_cleaning_recipe(session_id: str, request: FitRecipeRequest):
    """
    Split the dataset, compute all imputation statistics on the training split only,
    and return a serializable recipe dict.

    Strategy options: positional (default), random, time, stratified, group.
    Apply the recipe to train and test separately via POST /clean/apply to prevent
    data leakage — statistics are never recomputed on the test split.
    """
    try:
        filename, df = load_session(session_id)
    except FileNotFoundError:
        raise HTTPException(status_code=404, detail="Session not found")

    frac = max(0.5, min(0.95, request.train_fraction))
    train_df, test_df = split_dataframe(
        df,
        strategy=request.split_strategy,
        train_fraction=frac,
        random_state=request.random_state,
        time_col=request.time_col,
        stratify_col=request.stratify_col,
        group_col=request.group_col,
    )

    recipe = CleaningRecipe.fit(train_df, request.actions)
    recipe["split_metadata"] = {
        "strategy": request.split_strategy,
        "random_state": request.random_state,
        "time_col": request.time_col,
        "group_col": request.group_col,
        "stratify_col": request.stratify_col,
    }

    response: dict = {
        "recipe": recipe,
        "train_rows": len(train_df),
        "test_rows": len(test_df),
        "split_fraction": frac,
        "split_strategy_used": request.split_strategy,
        "note": (
            "Statistics fitted on training rows only. "
            "Call POST /clean/apply/{session_id} with this recipe on each split."
        ),
    }
    if request.split_strategy == "group" and request.group_col:
        response["train_groups"] = train_df[request.group_col].unique().to_list()
        response["test_groups"]  = test_df[request.group_col].unique().to_list()
    if request.split_strategy == "time":
        response["sort_column"] = request.time_col
    return response


@app.post("/clean/apply/{session_id}")
def apply_cleaning_recipe(session_id: str, request: ApplyRecipeRequest):
    """
    Apply a pre-fitted cleaning recipe (returned by /clean/fit) to the session
    data without recomputing any statistics.  Safe to call on both the train
    and test splits with the same recipe.
    """
    try:
        filename, df = load_session(session_id)
    except FileNotFoundError:
        raise HTTPException(status_code=404, detail="Session not found")

    recipe = request.recipe
    if not isinstance(recipe, dict) or "steps" not in recipe:
        raise HTTPException(status_code=422, detail="Invalid recipe: missing 'steps' key")

    backup_path = SESSION_DIR / session_id / "backup.parquet"
    try:
        df.write_parquet(backup_path)
    except Exception as e:
        print(f"Warning: Could not save backup: {e}")

    cleaned_df, changelog = CleaningRecipe.apply(df, recipe)
    save_session(session_id, filename, cleaned_df)

    if not session_id.isdigit():
        try:
            db.add_history(
                session_id, "clean",
                {"recipe_version": recipe.get("version"), "steps": len(recipe["steps"])},
                f"Applied pre-fitted recipe ({len(recipe['steps'])} steps): {', '.join(changelog[:3])}"
            )
            db.update_project(session_id, row_count=len(cleaned_df), column_count=len(cleaned_df.columns))
        except Exception:
            pass

    return {
        "status": "ok",
        "row_count": len(cleaned_df),
        "column_count": len(cleaned_df.columns),
        "changes": changelog,
        "recipe_version": recipe.get("version", 1),
        "fitted_at": recipe.get("fitted_at"),
        "can_undo": backup_path.exists(),
    }


# ── Gap 9: Schema contract / drift detection ─────────────────────────────────

class ValidateSchemaRequest(BaseModel):
    recipe: dict


@app.post("/clean/validate-schema/{session_id}")
def validate_schema(session_id: str, request: ValidateSchemaRequest):
    """
    Validate the current session data against the schema snapshot embedded in a
    CleaningRecipe (produced by POST /clean/fit).

    Returns violations in four categories:
    - missing_column (high): column present at training is absent now
    - dtype_changed (high): column dtype no longer matches training
    - new_categories (medium): unseen categorical values detected
    - range_breach / distribution_shift (medium): numeric values outside training bounds
    """
    try:
        filename, df = load_session(session_id)
    except FileNotFoundError:
        raise HTTPException(status_code=404, detail="Session not found")

    recipe = request.recipe
    snapshot = recipe.get("schema_snapshot")
    if not snapshot:
        raise HTTPException(
            status_code=422,
            detail=(
                "Recipe does not contain a 'schema_snapshot'. "
                "Re-fit the recipe via POST /clean/fit to include one."
            ),
        )

    violations = SchemaSnapshot.validate(df, snapshot)
    high = [v for v in violations if v.get("severity") == "high"]
    medium = [v for v in violations if v.get("severity") == "medium"]

    return {
        "session_id": session_id,
        "schema_ok": len(violations) == 0,
        "violation_count": len(violations),
        "high_severity": len(high),
        "medium_severity": len(medium),
        "violations": violations,
        "note": (
            "High-severity violations will likely cause the recipe to fail or produce "
            "wrong results. Medium-severity issues should be investigated before "
            "deploying the model."
        ),
    }


# ── Gap 13: Target-conditional missingness endpoint ─────────────────────────

class ClassifyMissingnessRequest(BaseModel):
    target: Optional[str] = None


@app.post("/clean/classify-missingness/{session_id}")
def classify_missingness_endpoint(session_id: str, request: ClassifyMissingnessRequest):
    """
    Classify the missingness mechanism for every column that has null values.

    Returns MCAR / MAR / MNAR / TARGET_LEAKAGE per column.
    Pass target= to enable the target-conditional check (Gap 13): if missingness
    correlates with the target column (|r| > 0.3), the column is flagged as
    TARGET_LEAKAGE — the most actionable classification, overriding MNAR/MAR/MCAR.
    """
    try:
        filename, df = load_session(session_id)
    except FileNotFoundError:
        raise HTTPException(status_code=404, detail="Session not found")

    if request.target and request.target not in df.columns:
        raise HTTPException(
            status_code=422,
            detail=f"Target column '{request.target}' not found in dataset.",
        )

    advisor = DataQualityAdvisor()
    classifications: dict[str, dict] = {}
    for col in df.columns:
        if df[col].null_count() > 0:
            mtype, reason = advisor.classify_missingness(col, df, request.target)
            classifications[col] = {"type": mtype, "reason": reason}

    return {
        "session_id": session_id,
        "target": request.target,
        "column_count": len(classifications),
        "classifications": classifications,
    }


# ── Gap 5: Target leakage detection ─────────────────────────────────────────

_LEAKAGE_KEYWORDS = [
    "total", "lifetime", "all_time", "alltime", "cumulative",
    "running", "ytd", "ever", "aggregate", "overall",
    "final", "end_", "last_", "trailing",
]

# Gap 7: keywords for transparent discount-imputation detection
_DISCOUNT_COL_KWS = {"discount", "sale", "promo", "reduced", "deal", "actual_price"}
_ORIGINAL_COL_KWS = {"original", "list", "mrp", "regular", "full", "base", "msrp"}


def detect_temporal_leakage(
    df: pl.DataFrame,
    target: str,
    time_col: Optional[str] = None,
) -> list[dict]:
    """Heuristic scan for features that are temporally leaky relative to *target*.

    Two signals are checked independently and combined:
    1. Keyword heuristic — column names containing aggregate/lifetime keywords.
    2. Near-perfect correlation with the target (|r| > 0.95) — likely a proxy.
    """
    _NUMERIC = (pl.Float64, pl.Float32, pl.Int64, pl.Int32, pl.Int16, pl.Int8)
    target_is_numeric = target in df.columns and df[target].dtype in _NUMERIC

    warnings: list[dict] = []

    for col in df.columns:
        if col in (target, time_col):
            continue

        col_lower = col.lower()
        reasons: list[str] = []

        # Signal 1: keyword match
        matched_kw = next((kw for kw in _LEAKAGE_KEYWORDS if kw in col_lower), None)
        if matched_kw:
            reasons.append(
                f"Column name contains temporal-aggregate keyword '{matched_kw}'. "
                f"This value may not exist at prediction time."
            )

        # Signal 2: near-perfect correlation with target
        if target_is_numeric and df[col].dtype in _NUMERIC:
            try:
                pdf = df[[col, target]].drop_nulls().to_pandas()
                r = pdf[col].corr(pdf[target])
                if pd.notna(r) and abs(r) > 0.95:
                    reasons.append(
                        f"Suspiciously high correlation with target '{target}' "
                        f"(r={r:+.2f}). May be derived from or identical to the target."
                    )
            except Exception:
                pass

        if reasons:
            warnings.append({
                "column": col,
                "leakage_risk": "high" if len(reasons) > 1 else "medium",
                "reasons": reasons,
                "suggested_action": "drop_column",
            })

    return warnings


class DetectLeakageRequest(BaseModel):
    target: str
    time_col: Optional[str] = None


@app.post("/clean/detect-leakage/{session_id}")
def detect_leakage(session_id: str, request: DetectLeakageRequest):
    """
    Scan the dataset for features that are likely to cause target leakage.

    Two heuristics are applied:
    - Column names containing temporal-aggregate keywords (total, lifetime, ytd, …)
    - Features that correlate ≥ 0.95 with the target (proxy / derived columns)

    Returns warnings with suggested actions. No data is modified.
    """
    try:
        filename, df = load_session(session_id)
    except FileNotFoundError:
        raise HTTPException(status_code=404, detail="Session not found")

    if request.target not in df.columns:
        raise HTTPException(
            status_code=422,
            detail=f"Target column '{request.target}' not found in dataset",
        )
    if request.time_col and request.time_col not in df.columns:
        raise HTTPException(
            status_code=422,
            detail=f"Time column '{request.time_col}' not found in dataset",
        )

    leakage_warnings = detect_temporal_leakage(df, request.target, request.time_col)

    return {
        "session_id": session_id,
        "target": request.target,
        "time_col": request.time_col,
        "warning_count": len(leakage_warnings),
        "leakage_warnings": leakage_warnings,
        "high_risk_columns": [w["column"] for w in leakage_warnings if w["leakage_risk"] == "high"],
        "note": (
            "Heuristic scan only — review each warning manually before dropping. "
            "Pass suggested_action='drop_column' items to /clean/{session_id} to remove them."
        ),
    }


# ── Gap 8: Near-duplicate detection (fuzzy matching) ─────────────────────────

def detect_near_duplicates(
    df: pl.DataFrame, col: str, threshold: int = 90
) -> list[dict]:
    """Pairwise fuzzy similarity on unique string values in *col*.

    Uses rapidfuzz if available, falls back to thefuzz, then difflib.SequenceMatcher.
    Caps at 500 unique values (O(n²) — ~125k comparisons) and returns at most 100 pairs.
    """
    try:
        from rapidfuzz import fuzz as _fuzz
        _ratio = lambda a, b: _fuzz.ratio(a, b)
    except ImportError:
        try:
            from thefuzz import fuzz as _fuzz  # type: ignore
            _ratio = lambda a, b: _fuzz.ratio(a, b)
        except ImportError:
            import difflib
            _ratio = lambda a, b: difflib.SequenceMatcher(None, a, b).ratio() * 100

    values = df[col].drop_nulls().unique().to_list()
    if len(values) > 500:
        values = values[:500]

    near_dupes: list[dict] = []
    for i in range(len(values)):
        a = str(values[i]).lower()
        for j in range(i + 1, len(values)):
            b = str(values[j]).lower()
            score = _ratio(a, b)
            if score >= threshold:
                near_dupes.append({
                    "value_a": values[i],
                    "value_b": values[j],
                    "similarity": round(float(score), 1),
                })

    near_dupes.sort(key=lambda x: x["similarity"], reverse=True)
    return near_dupes[:100]


class NearDuplicateRequest(BaseModel):
    column: str
    threshold: int = 90  # 0-100; lower = more pairs returned


@app.post("/clean/detect-near-duplicates/{session_id}")
def detect_near_duplicates_endpoint(session_id: str, request: NearDuplicateRequest):
    """
    Run pairwise fuzzy similarity on a string column's unique values and return
    pairs whose similarity score meets the threshold.

    Uses rapidfuzz > thefuzz > difflib (whichever is installed).
    Capped at 500 unique values for performance; returns top-100 pairs.
    """
    try:
        filename, df = load_session(session_id)
    except FileNotFoundError:
        raise HTTPException(status_code=404, detail="Session not found")

    if request.column not in df.columns:
        raise HTTPException(
            status_code=422,
            detail=f"Column '{request.column}' not found in dataset",
        )

    threshold = max(1, min(100, request.threshold))

    try:
        pairs = detect_near_duplicates(df, request.column, threshold)
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"Fuzzy match failed: {exc}")

    unique_count = df[request.column].drop_nulls().n_unique()
    capped = unique_count > 500

    return {
        "session_id": session_id,
        "column": request.column,
        "threshold": threshold,
        "unique_values_scanned": min(unique_count, 500),
        "unique_values_total": unique_count,
        "capped_at_500": capped,
        "pair_count": len(pairs),
        "near_duplicate_pairs": pairs,
        "note": (
            "Results are capped at 100 pairs sorted by similarity (descending). "
            "Values are compared lower-cased."
            + (" WARNING: Only the first 500 unique values were scanned." if capped else "")
        ),
    }


class EDAColumnStats(BaseModel):
    column: str
    dtype: str
    role: str = "unknown"  # identifier | numerical | categorical | temporal | binary | text
    mean: Optional[float] = None
    median: Optional[float] = None
    std: Optional[float] = None
    min_val: Optional[float] = None
    max_val: Optional[float] = None
    unique_count: int
    top_values: list[dict]  # For categorical columns

class EDAResponse(BaseModel):
    session_id: str
    numeric_columns: list[str]
    categorical_columns: list[str]
    date_columns: list[str]
    identifier_columns: list[str]  # NEW — excluded from analysis
    binary_columns: list[str]      # NEW
    column_stats: list[EDAColumnStats]
    correlation_matrix: dict  # column -> {column: correlation}
    insights: list[str]
    warnings: list[str]  # NEW — edge case warnings

@app.get("/eda/{session_id}", response_model=EDAResponse)
def get_eda(session_id: str):
    """
    Perform Exploratory Data Analysis.
    Uses the smart ColumnClassifier to properly detect and exclude ID columns.
    Results are cached — revisits are <5ms.
    """
    # ── Cache hit? ─────────────────────────────────────────────────
    cached = _cache.get(session_id, "eda_response")
    if cached is not None:
        return EDAResponse(**cached)

    try:
        filename, df = load_session(session_id)
    except FileNotFoundError:
        raise HTTPException(status_code=404, detail="Session not found")

    # ── Reuse cached profile if available ─────────────────────────
    profile = _cache.get(session_id, "profile")
    if profile is None:
        classifier = ColumnClassifier()
        profile = classifier.classify(df)
        _cache.put(session_id, "profile", profile)

    numeric_cols     = profile.numericals
    categorical_cols = profile.categoricals
    date_cols        = profile.temporals
    identifier_cols  = profile.identifiers
    binary_cols      = profile.binaries

    # Compute column statistics
    column_stats = []
    for col in df.columns:
        cp = profile.profiles.get(col)
        role = cp.role if cp else "unknown"

        stats = EDAColumnStats(
            column=col,
            dtype=str(df[col].dtype),
            role=role,
            unique_count=df[col].n_unique(),
            top_values=[]
        )

        if col in numeric_cols:
            col_data = df[col].drop_nulls()
            if len(col_data) > 0:
                stats.mean    = round(float(col_data.mean()), 2)
                stats.median  = round(float(col_data.median()), 2)
                stats.std     = round(float(col_data.std()), 2)
                stats.min_val = float(col_data.min())
                stats.max_val = float(col_data.max())

        if col in categorical_cols or col in binary_cols:
            stats.top_values = df[col].value_counts().head(5).to_dicts()

        column_stats.append(stats)

    # Correlation matrix — only non-identifier numerics
    correlation_matrix = {}
    analysis_numeric = [c for c in numeric_cols if c not in identifier_cols]
    if len(analysis_numeric) > 1:
        for col1 in analysis_numeric:
            correlation_matrix[col1] = {}
            for col2 in analysis_numeric:
                try:
                    corr = df.select(pl.corr(col1, col2)).item()
                    correlation_matrix[col1][col2] = round(float(corr), 3) if corr is not None else 0
                except Exception:
                    correlation_matrix[col1][col2] = 0

    # Smart insights (using the new engine's rule set)
    detector = AnomalyDetector()
    warnings = detector.detect(df, profile)

    insights = []
    if identifier_cols:
        insights.append(
            f"ID columns detected and excluded from analysis: {', '.join(identifier_cols)}."
        )

    for col1 in correlation_matrix:
        for col2, corr in correlation_matrix[col1].items():
            if col1 != col2 and abs(corr) > 0.7:
                insights.append(
                    f"Strong {'positive' if corr > 0 else 'negative'} correlation "
                    f"({corr:.2f}) between '{col1}' and '{col2}'."
                )

    for col in analysis_numeric[:3]:
        col_data = df[col].drop_nulls()
        if len(col_data) > 20:
            mean   = float(col_data.mean())
            median = float(col_data.median())
            std    = float(col_data.std()) or 1.0
            if abs(mean - median) / std > 0.5:
                direction = "right-skewed" if mean > median else "left-skewed"
                insights.append(
                    f"'{col}' is {direction} (mean: {mean:.2f}, median: {median:.2f})."
                )

    resp_data = {
        "session_id": session_id,
        "numeric_columns": numeric_cols,
        "categorical_columns": categorical_cols,
        "date_columns": date_cols,
        "identifier_columns": identifier_cols,
        "binary_columns": binary_cols,
        "column_stats": [s.model_dump() for s in column_stats],
        "correlation_matrix": correlation_matrix,
        "insights": insights[:10],
        "warnings": warnings[:5]
    }
    resp_data = make_json_safe(resp_data)
    _cache.put(session_id, "eda_response", resp_data)

    return EDAResponse(**resp_data)

# ============== PHASE 4: INTELLIGENCE LAYER ==============

class InsightCard(BaseModel):
    title: str
    description: str
    impact: str = "medium"   # "high" | "medium" | "low"
    recommendation: str = ""  # Actionable next step
    decision_implication: str = "" # Strategic next step
    confidence: str = "medium" # "high" | "medium" | "low"
    chart_type: str = "none"  # "bar" | "pie" | "line" | "scatter" | "none"
    chart_data: Optional[dict] = None
    # Keep legacy field for backward compat
    importance: str = "medium"

class RecommendationCard(BaseModel):
    priority: int
    action: str
    timeframe: str
    owner: str
    linked_insight: str = ""
    impact: str = "Medium"

class InsightsResponse(BaseModel):
    session_id: str
    executive_summary: str
    strategic_brief: list[InsightCard]
    recommendations: list[RecommendationCard]
    warnings: list[str] = []       # NEW — edge case / data quality alerts
    computed_metrics: Optional[dict] = None  # NEW — Revenue, Return Rate, AOV

@app.get("/insights/{session_id}", response_model=InsightsResponse)
def get_insights(session_id: str):
    """
    Generate smart business insights using the new insight engine.
    Results are cached — subsequent calls are <5ms.
    Falls back to background-job result if /analyze was used.
    """
    try:
        # ── Validate session exists first ─────────────────────────────
        if not session_exists(session_id):
            print(f"[ERROR] Session {session_id} not found")
            raise HTTPException(status_code=404, detail="Session not found or expired")
        
        # ── Check background job result first ─────────────────────────
        job = _jobs.get(session_id)
        if job and job.status == "done" and job.result:
            result = job.result
        else:
            # ── Check full result cache ──────────────────────────────
            cached_resp = _cache.get(session_id, "insight_response")
            if cached_resp is not None:
                print(f"[CACHE HIT] Returning cached insight response for {session_id}")
                return cached_resp

            result = _cache.get(session_id, "insight_result")

        if result is None:
            # ── Cold path: run synchronously with timeout ─────────────────────────
            print(f"[COLD PATH] Generating insights for session {session_id}")
            try:
                filename, df = load_session(session_id)
                print(f"[LOADED] Session {session_id}: {filename}, shape={df.shape}")
            except FileNotFoundError as e:
                print(f"[ERROR] Session file not found: {e}")
                raise HTTPException(status_code=404, detail="Session not found")
            except Exception as e:
                print(f"[ERROR] Failed to load session: {type(e).__name__}: {str(e)}")
                raise HTTPException(status_code=500, detail=f"Failed to load session data: {str(e)}")

            # Run insight engine with timeout protection
            import concurrent.futures
            with concurrent.futures.ThreadPoolExecutor(max_workers=1) as executor:
                future = executor.submit(run_insight_engine, df, 10, 0)
                try:
                    result = future.result(timeout=60)  # 60 second timeout
                    _cache.put(session_id, "insight_result", result)
                    print(f"[SUCCESS] Insights generated for {session_id}")
                except concurrent.futures.TimeoutError:
                    print(f"[ERROR] Insight generation timed out for session {session_id}")
                    raise HTTPException(
                        status_code=504,
                        detail="Insight generation is taking longer than expected. Please try using the 'Analyze' button for background processing."
                    )
                except Exception as e:
                    print(f"[ERROR] Insight engine failed: {type(e).__name__}: {str(e)}")
                    import traceback
                    traceback.print_exc()
                    raise

        # Validate result structure
        if not isinstance(result, dict):
            print(f"[ERROR] Invalid result type: {type(result)}")
            raise ValueError(f"Invalid insight result type: {type(result)}")
        
        if "strategic_brief" not in result:
            print(f"[ERROR] Missing strategic_brief in result. Keys: {result.keys()}")
            raise ValueError("Invalid insight result: missing strategic_brief")

        # Convert flat dicts → InsightCard objects
        insight_cards = []
        strategic_brief = result.get("strategic_brief", [])
        
        if not isinstance(strategic_brief, list):
            print(f"[ERROR] strategic_brief is not a list: {type(strategic_brief)}")
            strategic_brief = []
        
        for idx, ins in enumerate(strategic_brief):
            try:
                if not isinstance(ins, dict):
                    print(f"[WARNING] Insight {idx} is not a dict: {type(ins)}")
                    continue
                    
                card = InsightCard(
                    title=ins.get("title") or "",
                    description=ins.get("description") or "",
                    impact=ins.get("impact") or "medium",
                    importance=ins.get("impact") or "medium",
                    recommendation=ins.get("recommendation") or "",
                    decision_implication=ins.get("decision_implication") or "",
                    confidence=ins.get("confidence") or "medium",
                    chart_type=ins.get("chart_type") or "none",
                    chart_data=ins.get("chart_data"),
                )
                insight_cards.append(card)
            except Exception as e:
                print(f"[WARNING] Failed to convert insight {idx}: {type(e).__name__}: {str(e)}")
                continue

        exec_summary = result.get("executive_summary", "Analysis complete.")
        raw_recommendations = result.get("recommendations", [])
        warnings = result.get("warnings", [])
        computed_metrics = result.get("computed_metrics", {})
        
        # Validate and convert recommendations to proper format
        recommendations = []
        for idx, rec in enumerate(raw_recommendations):
            try:
                # If it's already a dict with required fields, use it
                if isinstance(rec, dict) and "action" in rec:
                    recommendations.append(rec)
                # If it's a string, convert to proper format
                elif isinstance(rec, str):
                    print(f"[WARNING] Recommendation {idx} is a string, converting to dict")
                    recommendations.append({
                        "priority": idx + 1,
                        "action": rec,
                        "timeframe": "Immediate",
                        "owner": "Team",
                        "linked_insight": "",
                        "impact": "High"
                    })
                else:
                    print(f"[WARNING] Skipping invalid recommendation {idx}: {type(rec)}")
            except Exception as e:
                print(f"[WARNING] Failed to process recommendation {idx}: {e}")
                continue

        # Persist to project DB (legacy system — skip for new integer IDs)
        if not session_id.isdigit():
            try:
                db.update_project(session_id, executive_summary=exec_summary,
                                  recommendations=json.dumps(recommendations[:5]))
                db.add_history(session_id, "insight", None,
                               f"Generated {len(insight_cards)} smart insights")
            except Exception as e:
                print(f"[WARNING] Failed to persist to DB: {e}")
                pass

        # Safe Return Layer (Step 1 - safe return layer)
        resp = InsightsResponse(
            session_id=session_id,
            executive_summary=exec_summary,
            strategic_brief=insight_cards,
            recommendations=recommendations,
            warnings=warnings,
            computed_metrics=computed_metrics,
        )
        
        # Final Validation Guard
        assert isinstance(resp.strategic_brief, list), "API Contract: strategic_brief must be a list"
        print(f"[SUCCESS] INSIGHTS OUTPUT: {len(resp.strategic_brief)} cards mapped for session {session_id}")
        
        _cache.put(session_id, "insight_response", resp)
        return resp
    except HTTPException:
        # Re-raise HTTP exceptions (404, 504, etc.)
        raise
    except Exception as e:
        # Log the full error for debugging
        import traceback
        print(f"[ERROR] Insights endpoint failed for session {session_id}:")
        print(f"[ERROR] {type(e).__name__}: {str(e)}")
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Failed to generate insights: {str(e)}")


# ============== KPI ENGINE ==============

KPI_KEYWORDS = [
    "revenue", "sales", "profit", "amount", "income", "price", "cost", "total", "quantity", "budget", "expense", "margin",
    "ltv", "lifetime", "burn", "runway"
]

class CustomKPIRequest(BaseModel):
    column: str
    aggregation: str  # sum | avg | growth
    label: str

@app.post("/kpis/{session_id}/custom")
def create_custom_kpi(session_id: str, body: CustomKPIRequest):
    """Create a user-defined KPI from a selected numeric column and aggregation."""
    try:
        _, df = load_session(session_id)
    except FileNotFoundError:
        raise HTTPException(status_code=404, detail="Session not found")

    if body.column not in df.columns:
        raise HTTPException(status_code=400, detail=f"Column '{body.column}' not found")

    # Use polars-friendly numeric check
    if not df[body.column].dtype.is_numeric():
        raise HTTPException(status_code=400, detail="Custom KPI requires a numeric column")

    aggregation = body.aggregation.lower().strip()
    if aggregation not in {"sum", "avg", "growth"}:
        raise HTTPException(status_code=400, detail="aggregation must be one of: sum, avg, growth")

    col_data = df[body.column].drop_nulls()
    if len(col_data) == 0:
        raise HTTPException(status_code=400, detail="Column has no non-null values")

    kpi = {
        "label": body.label.strip() or f"Custom {body.column}",
        "column": body.column,
        "avg": round(float(col_data.mean()), 2),
        "min": round(float(col_data.min()), 2),
        "max": round(float(col_data.max()), 2),
        "count": len(col_data),
    }

    if aggregation == "sum":
        metric_value = float(col_data.sum())
        kpi["value"] = metric_value
        kpi["formatted"] = _format_number(metric_value)
        kpi["trend"] = "flat"
        kpi["change_pct"] = 0
    elif aggregation == "avg":
        metric_value = float(col_data.mean())
        kpi["value"] = metric_value
        kpi["formatted"] = _format_number(metric_value)
        kpi["trend"] = "flat"
        kpi["change_pct"] = 0
    else:
        # Re-use logic for growth from profile if possible, or simple split
        metric_value = 0.0
        change_pct = 0.0
        # Try to find a temporal column
        profile = _cache.get(session_id, "profile")
        date_cols = profile.temporals if profile else []
        
        if date_cols:
            ordered = df.sort(date_cols[0])
            midpoint = len(ordered) // 2
            if midpoint > 0:
                first_half = float(ordered[:midpoint][body.column].drop_nulls().sum())
                second_half = float(ordered[midpoint:][body.column].drop_nulls().sum())
                metric_value = second_half
                if first_half != 0:
                    change_pct = round(((second_half - first_half) / abs(first_half)) * 100, 1)
        
        kpi["value"] = metric_value
        kpi["formatted"] = f"{change_pct:+.1f}%"
        kpi["change_pct"] = change_pct
        kpi["trend"] = "up" if change_pct > 2 else ("down" if change_pct < -2 else "flat")

    return {"session_id": session_id, "kpi": kpi}

def _format_number(val: float) -> str:
    """Format number as Indian Rupee with smart scaling."""
    abs_val = abs(val)
    sign = "" if val >= 0 else "-"
    if abs_val >= 1_00_00_000:      # 1 crore
        return f"{sign}₹{abs_val/1_00_00_000:.2f} Cr"
    if abs_val >= 1_00_000:          # 1 lakh
        return f"{sign}₹{abs_val/1_00_000:.2f} L"
    if abs_val >= 1_000:
        return f"{sign}₹{abs_val/1_000:.1f}K"
    return f"{sign}₹{abs_val:,.0f}"

@app.get("/kpis/{session_id}")
def get_kpis(session_id: str):
    """Auto-detect and compute KPI metrics. ID columns are excluded. Cached."""
    # ── Cache hit? ─────────────────────────────────────────────────
    cached = _cache.get(session_id, "kpi_response")
    if cached is not None:
        return cached

    try:
        filename, df = load_session(session_id)
    except FileNotFoundError:
        raise HTTPException(status_code=404, detail="Session not found")

    # Reuse cached profile
    profile = _cache.get(session_id, "profile")
    if profile is None:
        classifier = ColumnClassifier()
        profile = classifier.classify(df)
        _cache.put(session_id, "profile", profile)

    # Computed business metrics first (Revenue, Return Rate, AOV)
    computer = MetricComputer()
    computed = computer.compute(df, profile)

    kpis = []

    # Add computed metrics as KPI cards
    for key, metric in computed.items():
        kpis.append({
            "label": metric.name,
            "column": key,
            "value": metric.value,
            "formatted": metric.formatted,
            "avg": metric.value,
            "min": 0,
            "max": metric.value,
            "count": len(df),
            "trend": "flat",
            "change_pct": 0,
            "description": metric.description,
        })

    # Also add raw numeric KPIs (excluding identifiers)
    numeric_cols = [c for c in profile.numericals if c not in profile.identifiers]
    categorical_cols = profile.categoricals
    date_cols = profile.temporals

    # Keyword-match KPI columns (excluding IDs)
    kpi_cols = []
    for col in numeric_cols:
        col_lower = col.lower()
        if any(kw in col_lower for kw in KPI_KEYWORDS):
            # Skip if already covered by computed metrics
            already_covered = (
                col == profile.price_col or
                col == profile.qty_col or
                col == profile.revenue_col
            )
            if not already_covered:
                kpi_cols.append(col)

    if not kpi_cols and not computed:
        kpi_cols = numeric_cols[:3]

    for col in kpi_cols[:3]:  # Max 3 additional raw KPIs
        try:
            col_data = df[col].drop_nulls()
            if len(col_data) == 0:
                continue

            total   = float(col_data.sum())
            avg     = float(col_data.mean())
            min_val = float(col_data.min())
            max_val = float(col_data.max())

            kpi = {
                "label": f"Total {col}",
                "column": col,
                "value": total,
                "formatted": _format_number(total),
                "avg": round(avg, 2),
                "min": round(min_val, 2),
                "max": round(max_val, 2),
                "count": len(col_data),
            }

            # Growth trend via date column
            if date_cols:
                try:
                    date_col = date_cols[0]
                    sorted_df = df.sort(date_col)
                    n = len(sorted_df)
                    mid = n // 2
                    if mid > 0:
                        first_half  = float(sorted_df[:mid][col].drop_nulls().sum())
                        second_half = float(sorted_df[mid:][col].drop_nulls().sum())
                        if first_half != 0:
                            change_pct = round(((second_half - first_half) / abs(first_half)) * 100, 1)
                            kpi["change_pct"] = change_pct
                            kpi["trend"] = "up" if change_pct > 2 else ("down" if change_pct < -2 else "flat")
                        else:
                            kpi["change_pct"] = 0
                            kpi["trend"] = "flat"
                except Exception:
                    kpi["change_pct"] = 0
                    kpi["trend"] = "flat"
            else:
                kpi["trend"] = "flat"
                kpi["change_pct"] = 0

            # Best/worst category (using a non-ID categorical)
            if categorical_cols:
                try:
                    cat_col = profile.category_col or categorical_cols[0]
                    grouped = df.group_by(cat_col).agg(
                        pl.sum(col).alias("total")
                    ).sort("total", descending=True)
                    if len(grouped) >= 2:
                        best  = grouped.head(1)
                        worst = grouped.tail(1)
                        kpi["best_category"] = {
                            "name": str(best[cat_col].item()),
                            "value": float(best["total"].item())
                        }
                        kpi["worst_category"] = {
                            "name": str(worst[cat_col].item()),
                            "value": float(worst["total"].item())
                        }
                except Exception:
                    pass

            kpis.append(kpi)
        except Exception:
            continue

    resp = {"session_id": session_id, "kpis": kpis[:6], "currency": "INR"}
    _cache.put(session_id, "kpi_response", resp)
    return resp


# ============== BACKGROUND ANALYSIS (Layer 2) ==============

def _run_background_analysis(session_id: str) -> None:
    """Run the full insight engine in a background thread."""
    try:
        _jobs.update(session_id, status="running", progress=5)

        filename, df = load_session(session_id)
        _jobs.update(session_id, progress=10)

        def on_progress(stage: str, pct: int) -> None:
            _jobs.update(session_id, progress=pct)

        # Run the full pipeline with progress callbacks
        result = run_insight_engine(
            df,
            max_insights=10,
            max_charts=8,
            progress_callback=on_progress,
        )

        # Store in both the job tracker and the cache
        _cache.put(session_id, "insight_result", result)
        _jobs.update(session_id, status="done", progress=100, result=result)

    except Exception as e:
        _jobs.update(session_id, status="error", error=str(e))


@app.post("/analyze/{session_id}")
def start_analysis(session_id: str):
    """
    Kick off the full insight pipeline in a background thread.
    Returns 202 immediately; frontend polls /analyze-status for progress.
    """
    _cache.delete(session_id, "insight_result")
    _cache.delete(session_id, "insight_response")

    # Don't start if already running
    if _jobs.is_running(session_id):
        job = _jobs.get(session_id)
        return JSONResponse(
            status_code=202,
            content={
                "session_id": session_id,
                "status": job.status,
                "progress": job.progress,
                "message": "Analysis already in progress."
            }
        )

    # Check if result is already cached
    cached = _cache.get(session_id, "insight_result")
    if cached is not None:
        return JSONResponse(
            status_code=200,
            content={
                "session_id": session_id,
                "status": "done",
                "progress": 100,
                "message": "Results available (cached)."
            }
        )

    # Verify session exists before starting
    if not session_exists(session_id):
        raise HTTPException(status_code=404, detail="Session not found")

    # Create job and submit to thread pool
    _jobs.create(session_id)
    _thread_pool.submit(_run_background_analysis, session_id)

    return JSONResponse(
        status_code=202,
        content={
            "session_id": session_id,
            "status": "pending",
            "progress": 0,
            "message": "Analysis started. Poll /analyze-status for progress."
        }
    )


@app.get("/analyze-status/{session_id}")
def get_analysis_status(session_id: str):
    """
    Poll for background analysis progress.
    Returns status, progress %, and result when done.
    """
    job = _jobs.get(session_id)

    # If no job exists but cache has result, return done
    if job is None:
        cached = _cache.get(session_id, "insight_result")
        if cached is not None:
            return {
                "session_id": session_id,
                "status": "done",
                "progress": 100,
                "has_result": True,
            }
        return {
            "session_id": session_id,
            "status": "not_started",
            "progress": 0,
            "has_result": False,
        }

    response = {
        "session_id": session_id,
        "status": job.status,
        "progress": job.progress,
        "has_result": job.status == "done",
    }
    if job.status == "error":
        response["error"] = job.error

    return response


@app.get("/cache-status/{session_id}")
def get_cache_status(session_id: str):
    """Debug endpoint: show what's cached for a session."""
    return {
        "session_id": session_id,
        "has_profile": _cache.has(session_id, "profile"),
        "has_eda": _cache.has(session_id, "eda_response"),
        "has_insights": _cache.has(session_id, "insight_result"),
        "has_kpis": _cache.has(session_id, "kpi_response"),
        "cache_stats": _cache.stats(),
    }

# ============== AI CHART EXPLANATION ==============

class ExplainRequest(BaseModel):
    chart_id: str
    chart_type: str
    chart_title: str
    columns_used: List[str]
    data_summary: Optional[str] = None

@app.post("/explain-chart/{session_id}")
def explain_chart(session_id: str, request: ExplainRequest):
    """Use Gemini to explain patterns in a specific chart."""
    if not GENAI_API_KEY:
        return {
            "pattern": "AI explanation requires a Gemini API key.",
            "importance": "Configure GEMINI_API_KEY to enable chart explanations.",
            "business_reason": "",
            "risk_or_opportunity": ""
        }

    try:
        filename, df = load_session(session_id)
    except FileNotFoundError:
        raise HTTPException(status_code=404, detail="Session not found")

    # Build context from the data
    col_stats = []
    for col in request.columns_used:
        if col in df.columns:
            if df[col].dtype in [pl.Float64, pl.Float32, pl.Int64, pl.Int32]:
                stats = df[col].drop_nulls()
                col_stats.append(f"{col}: min={stats.min()}, max={stats.max()}, mean={round(float(stats.mean()), 2)}, median={round(float(stats.median()), 2)}")
            elif df[col].dtype == pl.Utf8:
                unique_count = df[col].n_unique()
                top_vals = df[col].value_counts().head(3)
                top_list = [str(row[col]) for row in top_vals.iter_rows(named=True)]
                col_stats.append(f"{col}: {unique_count} unique values, top: {', '.join(top_list)}")

    prompt = f"""You are a senior data analyst. Analyze this chart and provide insights.

Chart Type: {request.chart_type}
Chart Title: {request.chart_title}
Columns Used: {', '.join(request.columns_used)}
Dataset: {filename} ({len(df)} rows, {len(df.columns)} columns)
Column Statistics:
{chr(10).join(col_stats)}

{f'Additional context: {request.data_summary}' if request.data_summary else ''}

Provide a JSON response with exactly these 4 fields:
{{
    "pattern": "What pattern or trend does this chart reveal? Be specific with numbers.",
    "importance": "Why is this insight important for the business?",
    "business_reason": "What likely business factors explain this pattern?",
    "risk_or_opportunity": "What specific risk or opportunity does this suggest?"
}}

Be concise (2-3 sentences each). Use real numbers from the statistics."""

    try:
        model = genai.GenerativeModel("gemini-1.5-flash")
        response = model.generate_content(prompt)
        text = response.text.strip()

        # Parse JSON from response
        if "```json" in text:
            text = text.split("```json")[1].split("```")[0].strip()
        elif "```" in text:
            text = text.split("```")[1].split("```")[0].strip()

        result = json.loads(text)
        return {
            "pattern": result.get("pattern", ""),
            "importance": result.get("importance", ""),
            "business_reason": result.get("business_reason", ""),
            "risk_or_opportunity": result.get("risk_or_opportunity", ""),
        }
    except Exception as e:
        return {
            "pattern": f"Unable to generate explanation: {str(e)}",
            "importance": "",
            "business_reason": "",
            "risk_or_opportunity": ""
        }


# ============== PDF / HTML REPORT EXPORT ==============

from fastapi.responses import HTMLResponse
import base64

from fastapi import BackgroundTasks
import shutil
import tempfile

def cleanup_temp_dir(temp_dir: str, pdf_path: str):
    """Cleanup function to be run after FileResponse."""
    try:
        if os.path.exists(temp_dir):
            shutil.rmtree(temp_dir)
            print(f"Cleaned up temp directory: {temp_dir}")
    except Exception as e:
        print(f"Error during cleanup: {e}")

@app.get("/debug-columns/{session_id}")
async def debug_columns(session_id: str):
    """
    Returns the ColumnMap resolution for a session without generating a PDF.
    """
    try:
        filename, pl_df = load_session(session_id)
        df = pl_df.to_pandas()
        cm = ColumnMap(df)
        return JSONResponse({
            "all_columns":  list(df.columns),
            "dtypes":       {c: str(t) for c, t in df.dtypes.items()},
            "resolved": {
                "numeric":   cm.numeric,
                "numeric2":  cm.numeric2,
                "category":  cm.category,
                "region":    cm.region,
                "date":      cm.date,
                "label":     cm.label,
            }
        })
    except Exception as e:
        raise HTTPException(status_code=404, detail=str(e))

@app.get("/export-report/{session_id}")
async def export_report(session_id: str):
    """
    [DEPRECATED] This endpoint used the old Matplotlib generator.
    Please use the new 'Professional Export' button in the dashboard/insights UI.
    """
    raise HTTPException(
        status_code=410, 
        detail="This endpoint is deprecated. Use POST /export-dashboard-pdf instead."
    )



# ============== PHASE 5: ADVANCED VISUALIZATIONS ==============

class ChartData(BaseModel):
    chart_id: str
    chart_type: str
    title: str
    description: str
    plotly_json: dict
    columns_used: List[str]
    priority_score: float = 50.0  # Higher = more interesting (0-100)
    insight_reason: str = ""  # Why this chart was prioritized (for tooltip)
    interest_level: str = "standard"  # "high", "recommended", or "standard"

class VizResponse(BaseModel):
    session_id: str
    charts: List[ChartData]
    total_generated: int

def _normalize_query_list(value: Optional[List[str]]) -> Optional[List[str]]:
    """Normalize FastAPI Query defaults when endpoint functions are called directly."""
    if value is None:
        return None
    if isinstance(value, list):
        return value
    default = getattr(value, "default", None)
    return default if isinstance(default, list) or default is None else None

@app.get("/generate-viz/{session_id}")
def generate_visualizations(
    session_id: str, 
    max_charts: int = 10,
    groupby: Optional[str] = None,
    chart_types: Optional[List[str]] = Query(None),
    focus: Optional[str] = None,
    exclude_types: Optional[List[str]] = Query(None)
):
    """
    Auto-generate advanced interactive charts based on dataset characteristics.
    Returns Plotly JSON for frontend rendering with react-plotly.js.

    For unknown-domain datasets (HR, salary, etc.) that were analyzed by the LLM,
    returns the stored LLM charts directly so the UI matches the PDF export.
    Falls back to the rule-based engine for known domains (sports, entertainment).
    """
    # ── LLM chart override for unknown-domain sessions ────────────────────
    # If the session has stored LLM results with charts, return those directly.
    # This makes the UI show the same charts as the PDF export.
    if session_id.isdigit():
        try:
            import sqlite3 as _sq, json as _jmod
            from pathlib import Path as _P
            _db_path = _P(__file__).parent / "data" / "insightstream.db"
            _conn = _sq.connect(str(_db_path))
            _row = _conn.execute(
                "SELECT llm_results_json FROM analysis_sessions WHERE id = ?",
                (int(session_id),)
            ).fetchone()
            _conn.close()

            if _row and _row[0]:
                _llm = _jmod.loads(_row[0])
                _stored = _llm.get("charts", [])
                if _stored:
                    # Convert stored plotly_json strings to ChartData objects
                    # using the same shape _internal_gen_viz returns
                    import plotly.io as _pio
                    _chart_list = []
                    for _i, _ch in enumerate(_stored):
                        _pj = _ch.get("plotly_json", "")
                        if not _pj:
                            continue
                        try:
                            # Strip template to avoid version-mismatch errors
                            _raw = _jmod.loads(_pj) if isinstance(_pj, str) else _pj
                            _raw.get("layout", {}).pop("template", None)
                            # plotly_json must be a dict (not a string) per ChartData schema
                            _plotly_dict = _raw
                            _chart_list.append(ChartData(
                                chart_id=_ch.get("id", f"llm_chart_{_i}"),
                                chart_type="bar",
                                title=_ch.get("title", f"Chart {_i+1}"),
                                description=_ch.get("title", ""),   # required field
                                plotly_json=_plotly_dict,            # dict, not string
                                columns_used=[],                     # required field
                                priority_score=90.0,
                                insight_reason="LLM-generated chart",
                                interest_level="high",
                            ))
                        except Exception as _ce:
                            print(f"[generate-viz] LLM chart {_i} parse failed: {_ce}")

                    if _chart_list:
                        print(f"[generate-viz] Returning {len(_chart_list)} LLM charts for session {session_id}")
                        # Return a VizResponse-compatible object
                        return VizResponse(
                            session_id=str(session_id),
                            charts=_chart_list,
                            total_generated=len(_chart_list),
                        )
        except Exception as _oe:
            print(f"[generate-viz] LLM override check failed: {_oe}")
    # ── END LLM override ──────────────────────────────────────────────────

    try:
        resp = _internal_gen_viz(
            session_id=session_id,
            max_charts=max_charts,
            groupby=groupby,
            chart_types=_normalize_query_list(chart_types),
            focus=focus,
            exclude_types=_normalize_query_list(exclude_types)
        )
        if not session_id.isdigit():
            try:
                db.add_history(
                    session_id,
                    "generate_viz",
                    {"groupby": groupby, "chart_types": chart_types},
                    f"Generated {resp.total_generated} visualizations"
                )
            except Exception as e:
                print(f"Warning: Could not log history: {e}")
        return resp
    except Exception as e:
        import traceback
        trace = traceback.format_exc()
        print(f"CRITICAL VIZ ERROR: {e}\n{trace}")
        raise HTTPException(status_code=500, detail=str(e))

def _internal_gen_viz(
    session_id: str, 
    max_charts: int = 10,
    groupby: Optional[str] = None,
    chart_types: Optional[List[str]] = None,
    focus: Optional[str] = None,
    exclude_types: Optional[List[str]] = None
):
    # Helper for type filtering
    def is_chart_allowed(ctype):
        if exclude_types and ctype in exclude_types: return False
        if chart_types and ctype not in chart_types: return False
        return True
    

    
    # Limit max_charts to 20 for performance
    if max_charts > 20: 
        max_charts = 20
        
    try:
        filename, df = load_session(session_id)
        print(f"Loaded session {session_id} for viz generation.")
        
        # Convert to pandas for Plotly
        pdf = df.to_pandas()
        charts = []
        

        
    except FileNotFoundError:
        print(f"Session {session_id} not found.")
        raise HTTPException(status_code=404, detail="Session not found")
    except Exception as e:
        import traceback
        print(f"Error in generate_visualizations: {e}")
        print(traceback.format_exc())
        raise HTTPException(status_code=500, detail=f"Backend Error: {str(e)}")
    
    # ── Smart column classification — exclude identifiers from all charts ──
    _classifier = ColumnClassifier()
    _profile = _classifier.classify(df)
    _id_cols = set(_profile.identifiers)

    # If no custom filters requested, delegate to SmartChartRecommender
    _use_smart = not groupby and not chart_types and not focus and not exclude_types
    if _use_smart:
        try:
            from insight_engine import detect_domain
            _domain_id = detect_domain(df)
            rec = SmartChartRecommender()
            smart_charts = rec.recommend(df, _profile, [],
                                         max_charts=max_charts,
                                         domain_id=_domain_id)
            return VizResponse(
                session_id=session_id,
                charts=[ChartData(**c) for c in smart_charts],
                total_generated=len(smart_charts)
            )
        except Exception as _e:
            print(f"SmartChartRecommender failed, falling back: {_e}")
            # Fall through to legacy path

    # Categorize columns — excluding identifiers
    numeric_cols = [c for c in pdf.columns
                    if pdf[c].dtype in ['int64', 'float64', 'int32', 'float32']
                    and c not in _id_cols]
    categorical_cols = [c for c in pdf.columns
                        if (pdf[c].dtype == 'object' or pdf[c].nunique() < 10)
                        and c not in _id_cols]
    
    # ============== SMART COLUMN SCORING ==============
    # Analyze data patterns to prioritize most interesting visualizations
    
    column_scores = {}
    column_insights = {}  # Track why each column is interesting
    
    def get_interest_level(score: float) -> str:
        """Convert priority score to interest level label."""
        if score >= 70:
            return "high"
        elif score >= 55:
            return "recommended"
        return "standard"
    
    # Score numeric columns by: variance, missing %, unique ratio
    for col in numeric_cols:
        score = 50
        reasons = []
        missing_pct = pdf[col].isna().sum() / len(pdf) * 100
        if missing_pct > 10:
            score += 15  # High missing = interesting distribution question
            reasons.append(f"{missing_pct:.0f}% missing values")
        try:
            cv = pdf[col].std() / (abs(pdf[col].mean()) + 0.001)
            if cv > 0.5:
                score += 10  # High coefficient of variation = spread worth showing
                reasons.append("High variance")
        except:
            pass
        column_scores[col] = min(score, 100)
        column_insights[col] = reasons if reasons else ["Standard distribution"]
    
    # Score categorical columns by: cardinality (3-8 ideal), balance
    for col in categorical_cols:
        score = 50
        reasons = []
        n_unique = pdf[col].nunique()
        if 3 <= n_unique <= 8:
            score += 20  # Ideal for pie/bar
            reasons.append(f"{n_unique} categories (ideal)")
        elif n_unique <= 2:
            score -= 10  # Binary less interesting
        else:
            reasons.append(f"{n_unique} categories")
        # Check balance (entropy-like)
        value_counts = pdf[col].value_counts(normalize=True)
        if value_counts.max() < 0.7:  # Not dominated by one value
            score += 10
            reasons.append("Balanced distribution")
        column_scores[col] = min(score, 100)
        column_insights[col] = reasons if reasons else ["Categorical variable"]
    
    # Find strong correlations for scatter prioritization
    strong_correlations = []
    if len(numeric_cols) >= 2:
        try:
            corr_matrix = pdf[numeric_cols].corr()
            for i, col1 in enumerate(numeric_cols):
                for col2 in numeric_cols[i+1:]:
                    corr = abs(corr_matrix.loc[col1, col2])
                    if corr > 0.5:
                        strong_correlations.append((col1, col2, corr))
            strong_correlations.sort(key=lambda x: x[2], reverse=True)
        except:
            pass
    
    # Sort columns by interest score for chart generation
    numeric_cols = sorted(numeric_cols, key=lambda c: column_scores.get(c, 50), reverse=True)
    categorical_cols = sorted(categorical_cols, key=lambda c: column_scores.get(c, 50), reverse=True)
    

    
    # Override primary category if groupby provided
    if groupby:
        # Boost score of groupby column if it's already categorical
        if groupby in categorical_cols:
            categorical_cols.remove(groupby)
            categorical_cols.insert(0, groupby)
            column_scores[groupby] = 100
        elif groupby in numeric_cols:
            # If it was numeric, treat it as categorical for grouping purposes
            numeric_cols.remove(groupby)
            categorical_cols.insert(0, groupby)
            column_scores[groupby] = 100

    # ============== PHASE 0: CORE BASICS ==============
    
    # 0.1 HISTOGRAM: Distribution of numeric columns (with optional color grouping)
    if is_chart_allowed("histogram"):
        # If groupby exists, use it for color
        force_color = groupby if groupby in categorical_cols else None
        
        for col in numeric_cols[:2]:  # Limit to first 2 numeric columns
            if pdf[col].nunique() > 5 and len(charts) < max_charts:
                try:
                    # Add color grouping if good categorical exists
                    color_col = None
                    if categorical_cols and pdf[categorical_cols[0]].nunique() <= 4:
                        color_col = categorical_cols[0]
                    
                    fig = px.histogram(
                        pdf, x=col,
                        color=color_col,
                        title=f"Distribution of {col}" + (f" by {color_col}" if color_col else ""),
                        marginal="rug",
                        barmode="overlay" if color_col else "relative",
                        opacity=0.7 if color_col else 1.0
                    )
                    fig.update_layout(template="plotly_dark", showlegend=bool(color_col))
                    if not color_col:
                        fig.update_traces(marker_color='#6366f1')
                    score = column_scores.get(col, 50) + (10 if color_col else 0)
                    reasons = column_insights.get(col, [])
                    charts.append(ChartData(
                        chart_id=f"histogram_{col}",
                        chart_type="histogram",
                        title=f"{col} Distribution",
                    description=f"Frequency distribution" + (f" colored by {color_col}" if color_col else ""),
                    plotly_json=json.loads(fig.update_layout(**CHART_LAYOUT_BASE).to_json()),
                    columns_used=[col] + ([color_col] if color_col else []),
                    priority_score=score,
                    insight_reason=" • ".join(reasons) if reasons else "Distribution analysis",
                    interest_level=get_interest_level(score)
                ))
                except:
                    pass
    


    # 0.2 BAR CHART: Categorical vs Numeric (median for robustness, horizontal for long labels)
    if is_chart_allowed("bar") and len(categorical_cols) >= 1 and len(numeric_cols) >= 1 and len(charts) < max_charts:
        try:
            cat_col = categorical_cols[0]
            num_col = numeric_cols[0]
            n_categories = pdf[cat_col].nunique()
            if n_categories <= 12 and cat_col != num_col:
                # Use median for robustness against outliers
                agg_df = pdf.groupby(cat_col)[num_col].median().reset_index()
                agg_df.columns = [cat_col, f'Median {num_col}']
                agg_df = agg_df.sort_values(f'Median {num_col}', ascending=True)
                
                # Zero-variance suppression
                if not is_chart_informative(agg_df[f'Median {num_col}'].tolist()):
                    print(f"[CHART SUPPRESSED] {num_col} by {cat_col} — all values flat")
                else:
                    # Use horizontal for many categories or long labels
                    use_horizontal = n_categories > 6 or pdf[cat_col].astype(str).str.len().max() > 10
                    
                    fig = px.bar(
                        agg_df, 
                        x=f'Median {num_col}' if use_horizontal else cat_col,
                        y=cat_col if use_horizontal else f'Median {num_col}',
                        orientation='h' if use_horizontal else 'v',
                        title=f"Median {num_col} by {cat_col}",
                        text_auto='.1f',
                        color=f'Median {num_col}',
                        color_continuous_scale='Viridis'
                    )
                    fig.update_layout(template="plotly_dark", showlegend=False, coloraxis_showscale=False)
                    score = (column_scores.get(cat_col, 50) + column_scores.get(num_col, 50)) / 2
                    reasons = column_insights.get(cat_col, []) + column_insights.get(num_col, [])
                    charts.append(ChartData(
                        chart_id=f"bar_{num_col}_{cat_col}",
                        chart_type="bar",
                        title=f"{num_col} by {cat_col}",
                        description=f"Median {num_col} comparison (robust to outliers)",
                        plotly_json=json.loads(fig.update_layout(**CHART_LAYOUT_BASE).to_json()),
                        columns_used=[cat_col, num_col],
                        priority_score=score,
                        insight_reason=" • ".join(reasons[:2]) if reasons else "Category comparison",
                        interest_level=get_interest_level(score)
                    ))

        except Exception as e:
            print(f"Bar Chart Error: {e}")
    


    # 0.3 PIE CHART: With auto-collapse small slices into "Other"
    if is_chart_allowed("pie") and len(categorical_cols) >= 1 and len(charts) < max_charts:
        cat_col = categorical_cols[0]
        n_unique = pdf[cat_col].nunique()
        if n_unique <= 15:  # Allow more but collapse
            try:
                counts = pdf[cat_col].value_counts().reset_index()
                counts.columns = [cat_col, 'count']
                
                # Auto-collapse slices < 4% into "Other"
                total = counts['count'].sum()
                threshold = 0.04 * total
                
                # Mark small categories as "Other"
                counts['category'] = counts.apply(
                    lambda row: row[cat_col] if row['count'] >= threshold else 'Other',
                    axis=1
                )
                counts = counts.groupby('category')['count'].sum().reset_index()
                counts = counts.sort_values('count', ascending=False)
                
                # Limit to 8 slices max
                if len(counts) > 8:
                    top_7 = counts.head(7)
                    other_sum = counts.iloc[7:]['count'].sum()
                    other_row = pd.DataFrame({'category': ['Other'], 'count': [other_sum]})
                    counts = pd.concat([top_7, other_row], ignore_index=True)
                
                fig = px.pie(
                    counts, names='category', values='count',
                    title=f"Proportion of {cat_col}",
                    hole=0.4,  # Slightly larger hole for modern look
                    color_discrete_sequence=px.colors.qualitative.Set2
                )
                fig.update_layout(template="plotly_dark")
                fig.update_traces(textposition='inside', textinfo='percent+label', 
                                  pull=[0.02] * len(counts))  # Slight pull for emphasis
                charts.append(ChartData(
                    chart_id=f"pie_{cat_col}",
                    chart_type="pie",
                    title=f"{cat_col} Breakdown",
                    description=f"Percentage distribution (small slices grouped as 'Other')",
                    plotly_json=json.loads(fig.update_layout(**CHART_LAYOUT_BASE).to_json()),
                    columns_used=[cat_col],
                    priority_score=column_scores.get(cat_col, 50) + 5  # Pie charts popular
                ))
            except:
                pass
    
    # 0.4 COUNT BAR: Stacked categorical frequency
    if is_chart_allowed("bar") and len(categorical_cols) >= 2 and len(charts) < max_charts:
        try:
            cat1, cat2 = categorical_cols[0], categorical_cols[1]
            if cat1 != cat2 and pdf[cat1].nunique() <= 8 and pdf[cat2].nunique() <= 6:
                fig = px.histogram(
                    pdf, x=cat1, color=cat2,
                    title=f"{cat1} Counts by {cat2}",
                    barmode="stack"  # Stacked shows proportions better
                )
                fig.update_layout(template="plotly_dark")
                charts.append(ChartData(
                    chart_id=f"countbar_{cat1}_{cat2}",
                    chart_type="count_bar",
                    title=f"{cat1} by {cat2}",
                    description=f"Frequency counts of {cat1} grouped by {cat2}",
                    plotly_json=json.loads(fig.update_layout(**CHART_LAYOUT_BASE).to_json()),
                    columns_used=[cat1, cat2]
                ))
        except Exception as e:
            print(f"Count Bar Error: {e}")
    
    # ============== PHASE 1: STATISTICAL & DISTRIBUTION ==============
    
    # Find best numeric column (highest variance or missing %)
    best_num = None
    best_num_score = 0
    for col in numeric_cols:
        try:
            variance = pdf[col].var() or 0
            missing = pdf[col].isna().sum() / len(pdf)
            score = variance / (pdf[col].mean()**2 + 0.001) + missing * 50  # CV + missing bonus
            if score > best_num_score:
                best_num_score = score
                best_num = col
        except:
            pass
    if not best_num and numeric_cols:
        best_num = numeric_cols[0]
    
    # Find best categorical (3-8 unique values, balanced)
    best_cat = None
    best_cat_score = 0
    for col in categorical_cols:
        n_unique = pdf[col].nunique()
        if 3 <= n_unique <= 8:
            balance = 1 - pdf[col].value_counts(normalize=True).max()
            score = balance * 100 + (8 - abs(n_unique - 5)) * 5
            if score > best_cat_score:
                best_cat_score = score
                best_cat = col
    if not best_cat and categorical_cols:
        best_cat = categorical_cols[0]
    
    # Force override again just in case loop reset it (though we set it before)
    if groupby and groupby in categorical_cols:
        best_cat = groupby
    
    # 1.1 VIOLIN PLOT: Best numeric by best categorical
    if is_chart_allowed("violin") and best_num and best_cat and len(charts) < max_charts:
        try:
            fig = px.violin(
                pdf, x=best_cat, y=best_num,
                box=True, points="all",
                color=best_cat,
                title=f"Distribution of {best_num} by {best_cat}",
                color_discrete_sequence=px.colors.qualitative.Set2
            )
            fig.update_layout(template="plotly_dark", showlegend=False)
            score = column_scores.get(best_num, 50) + 25  # Violin boost
            reasons = column_insights.get(best_num, []) + ["Shows distribution shape + outliers"]
            charts.append(ChartData(
                chart_id=f"violin_{best_num}_{best_cat}",
                chart_type="violin",
                title=f"{best_num} by {best_cat}",
                description=f"Full distribution with box stats and individual points",
                plotly_json=json.loads(fig.update_layout(**CHART_LAYOUT_BASE).to_json()),
                columns_used=[best_num, best_cat],
                priority_score=score,
                insight_reason=" • ".join(reasons[:3]),
                interest_level=get_interest_level(score)
            ))
        except:
            pass
    
    # 1.2 SCATTER + MARGINALS: Strongest correlation pair
    if is_chart_allowed("scatter") and len(strong_correlations) > 0 and len(charts) < max_charts:
        # Use the strongest correlation pair
        col1, col2, corr_val = strong_correlations[0]
        color_col = best_cat if best_cat and pdf[best_cat].nunique() <= 4 else None
        try:
            fig = px.scatter(
                pdf, x=col1, y=col2,
                color=color_col,
                marginal_x="histogram", marginal_y="violin",
                title=f"{col1} vs {col2} (r = {corr_val:.2f})",
                opacity=0.7,
                trendline="ols" if not color_col else None  # Add OLS trendline when no color grouping
            )
            fig.update_layout(template="plotly_dark")
            fig.update_traces(marker=dict(size=6), selector=dict(mode="markers"))
            # Style the trendline (if present)
            fig.update_traces(line=dict(color="#ef4444", dash="dash", width=2), selector=dict(mode="lines"))
            score = 60 + corr_val * 40  # Strong boost for high correlation
            charts.append(ChartData(
                chart_id=f"scatter_corr_{col1}_{col2}",
                chart_type="scatter_marginals",
                title=f"{col1} vs {col2}",
                description=f"Bivariate relationship with marginal distributions",
                plotly_json=json.loads(fig.update_layout(**CHART_LAYOUT_BASE).to_json()),
                columns_used=[col1, col2] + ([color_col] if color_col else []),
                priority_score=score,
                insight_reason=f"Strongest correlation (r = {corr_val:.2f})",
                interest_level=get_interest_level(score)
            ))
        except:
            pass
    elif is_chart_allowed("scatter") and len(numeric_cols) >= 2 and len(charts) < max_charts:
        # Fallback: first two numerics
        col1, col2 = numeric_cols[0], numeric_cols[1]
        try:
            fig = px.scatter(
                pdf, x=col1, y=col2,
                marginal_x="histogram", marginal_y="violin",
                title=f"{col1} vs {col2} with Distributions"
            )
            fig.update_layout(template="plotly_dark")
            charts.append(ChartData(
                chart_id=f"scatter_marginal_{col1}_{col2}",
                chart_type="scatter_marginals",
                title=f"{col1} vs {col2}",
                description=f"Bivariate analysis with marginal distributions",
                plotly_json=json.loads(fig.update_layout(**CHART_LAYOUT_BASE).to_json()),
                columns_used=[col1, col2],
                priority_score=55,
                insight_reason="Numeric relationship exploration",
                interest_level="recommended"
            ))
        except:
            pass
    
    # 1.3 BOX PLOT: With notches and points
    if is_chart_allowed("box") and best_num and best_cat and len(charts) < max_charts:
        try:
            fig = px.box(
                pdf, x=best_cat, y=best_num,
                notched=True, points="all",
                color=best_cat,
                title=f"Box Summary of {best_num} by {best_cat}",
                color_discrete_sequence=px.colors.qualitative.Pastel
            )
            fig.update_layout(template="plotly_dark", showlegend=False)
            score = column_scores.get(best_num, 50) + 15  # Box plot boost
            charts.append(ChartData(
                chart_id=f"box_{best_num}_{best_cat}",
                chart_type="box",
                title=f"{best_num} Summary",
                description=f"Quartiles, median, and confidence intervals with data points",
                plotly_json=json.loads(fig.update_layout(**CHART_LAYOUT_BASE).to_json()),
                columns_used=[best_num, best_cat],
                priority_score=score,
                insight_reason="Statistical summary with confidence notches",
                interest_level=get_interest_level(score)
            ))
        except:
            pass
    
    # 1.4 DENSITY HEATMAP: Joint distribution of top correlation pair
    if is_chart_allowed("density_heatmap") and len(numeric_cols) >= 2 and len(charts) < max_charts:
        # Use correlation pair if available, else first two
        if strong_correlations:
            col1, col2, _ = strong_correlations[0]
        else:
            col1, col2 = numeric_cols[0], numeric_cols[1]
        try:
            fig = px.density_heatmap(
                pdf, x=col1, y=col2,
                title=f"Density: {col1} vs {col2}",
                color_continuous_scale="Viridis",
                marginal_x="histogram", marginal_y="histogram"
            )
            fig.update_layout(template="plotly_dark")
            charts.append(ChartData(
                chart_id=f"density_{col1}_{col2}",
                chart_type="density_heatmap",
                title=f"Joint Distribution",
                description=f"Density concentration showing where data clusters",
                plotly_json=json.loads(fig.update_layout(**CHART_LAYOUT_BASE).to_json()),
                columns_used=[col1, col2],
                priority_score=58,
                insight_reason="Shows concentration patterns in joint distribution",
                interest_level="recommended"
            ))
        except:
            pass
    
    # ============== PHASE 2: HIERARCHICAL & SPECIALIZED ==============
    
    # Smart path detection for hierarchical charts
    path_candidates = []
    for col in categorical_cols:
        n_unique = pdf[col].nunique()
        if 3 <= n_unique <= 12:
            balance = 1 - pdf[col].value_counts(normalize=True).max()
            avg_label_len = pdf[col].astype(str).str.len().mean()
            score = balance * 50 + (12 - abs(n_unique - 6)) * 5
            path_candidates.append((col, n_unique, balance, avg_label_len, score))
    
    # Sort by score (best first)
    path_candidates.sort(key=lambda x: x[4], reverse=True)
    
    # 2.1 TREEMAP: Hierarchical categories (prefer when labels are long)
    if is_chart_allowed("treemap") and len(path_candidates) >= 2 and len(charts) < max_charts:
        path_cols = [p[0] for p in path_candidates[:3]]  # Top 2-3 columns
        val_col = best_num if best_num else None
        avg_label_len = sum(p[3] for p in path_candidates[:2]) / 2
        
        # Use treemap when labels are longer
        if avg_label_len > 6:
            try:
                if val_col:
                    treemap_df = pdf.dropna(subset=path_cols + [val_col])
                    fig = px.treemap(
                        treemap_df,
                        path=path_cols[:2],
                        values=val_col,
                        title=f"Breakdown of {val_col} by {' → '.join(path_cols[:2])}",
                        color_continuous_scale="Viridis"
                    )
                else:
                    treemap_df = pdf.dropna(subset=path_cols)
                    treemap_df['_count'] = 1
                    fig = px.treemap(
                        treemap_df,
                        path=path_cols[:2],
                        values='_count',
                        title=f"Hierarchy: {' → '.join(path_cols[:2])}"
                    )
                fig.update_layout(template="plotly_dark")
                depth = len(path_cols[:2])
                score = 60 + (depth * 15)  # Boost for depth
                charts.append(ChartData(
                    chart_id=f"treemap_{'_'.join(path_cols[:2])}",
                    chart_type="treemap",
                    title=f"Hierarchical Breakdown",
                    description=f"Part-to-whole analysis across {depth} levels",
                    plotly_json=json.loads(fig.update_layout(**CHART_LAYOUT_BASE).to_json()),
                    columns_used=path_cols[:2] + ([val_col] if val_col else []),
                    priority_score=score,
                    insight_reason=f"{depth}-level hierarchy detected • Good cardinality",
                    interest_level=get_interest_level(score)
                ))
            except:
                pass
    
    # 2.2 SUNBURST: Radial hierarchy (prefer when labels are short)
    if is_chart_allowed("sunburst") and len(path_candidates) >= 2 and len(charts) < max_charts:
        path_cols = [p[0] for p in path_candidates[:3]]
        val_col = best_num if best_num else None
        avg_label_len = sum(p[3] for p in path_candidates[:2]) / 2
        
        # Use sunburst when labels are shorter (fits better radially)
        if avg_label_len <= 10:
            try:
                if val_col:
                    sunburst_df = pdf.dropna(subset=path_cols[:2] + [val_col])
                    fig = px.sunburst(
                        sunburst_df,
                        path=path_cols[:2],
                        values=val_col,
                        title=f"Sunburst: {' → '.join(path_cols[:2])}",
                        color_continuous_scale="Viridis"
                    )
                else:
                    sunburst_df = pdf.dropna(subset=path_cols[:2])
                    sunburst_df['_count'] = 1
                    fig = px.sunburst(
                        sunburst_df,
                        path=path_cols[:2],
                        values='_count',
                        title=f"Sunburst: {' → '.join(path_cols[:2])}"
                    )
                fig.update_layout(template="plotly_dark")
                score = 58 + (len(path_cols[:2]) * 12)
                charts.append(ChartData(
                    chart_id=f"sunburst_{'_'.join(path_cols[:2])}",
                    chart_type="sunburst",
                    title=f"Radial Hierarchy",
                    description=f"Nested breakdown in radial format",
                    plotly_json=json.loads(fig.update_layout(**CHART_LAYOUT_BASE).to_json()),
                    columns_used=path_cols[:2] + ([val_col] if val_col else []),
                    priority_score=score,
                    insight_reason="Compact labels suit radial display",
                    interest_level=get_interest_level(score)
                ))
            except:
                pass
    
    # 2.3 FUNNEL CHART: Detect stage-like columns
    stage_keywords = ['stage', 'step', 'phase', 'funnel', 'conversion', 'status', 'level', 'tier']
    ordinal_patterns = ['low', 'medium', 'high', 'small', 'large', 'start', 'end', 'begin', 'complete']
    
    funnel_col = None
    funnel_score = 0
    for col in categorical_cols:
        col_lower = col.lower()
        # Check if column name matches stage keywords
        if any(kw in col_lower for kw in stage_keywords):
            funnel_col = col
            funnel_score = 85  # High priority for explicit stage columns
            break
        # Check if values look ordinal
        values = pdf[col].dropna().astype(str).str.lower().unique()
        if len(values) >= 2 and len(values) <= 8:
            if any(p in ' '.join(values) for p in ordinal_patterns):
                funnel_col = col
                funnel_score = 70
    
    if is_chart_allowed("funnel") and funnel_col and len(charts) < max_charts:
        try:
            funnel_data = pdf[funnel_col].value_counts().reset_index()
            funnel_data.columns = [funnel_col, 'count']
            # Sort by count descending for proper funnel shape
            funnel_data = funnel_data.sort_values('count', ascending=False)
            
            fig = px.funnel(
                funnel_data,
                x='count',
                y=funnel_col,
                title=f"Funnel: {funnel_col}",
                color_discrete_sequence=['#6366f1']
            )
            fig.update_layout(template="plotly_dark")
            charts.append(ChartData(
                chart_id=f"funnel_{funnel_col}",
                chart_type="funnel",
                title=f"{funnel_col} Funnel",
                description="Stage-by-stage breakdown showing progression",
                plotly_json=json.loads(fig.update_layout(**CHART_LAYOUT_BASE).to_json()),
                columns_used=[funnel_col],
                priority_score=funnel_score,
                insight_reason="Stage-like column detected" if funnel_score > 75 else "Ordinal values suggest progression",
                interest_level=get_interest_level(funnel_score)
            ))
        except:
            pass
    
    # 2.4 ICICLE: For deep hierarchies (3+ levels)
    if is_chart_allowed("icicle") and len(path_candidates) >= 3 and len(charts) < max_charts:
        path_cols = [p[0] for p in path_candidates[:3]]
        val_col = best_num if best_num else None
        
        try:
            # Limit to 2 levels max, and filter to top segments per level (Fix 6)
            cols_to_use = path_cols[:2]
            
            if val_col:
                icicle_df = pdf.dropna(subset=cols_to_use + [val_col])
                fig = px.icicle(
                    icicle_df,
                    path=cols_to_use,
                    values=val_col,
                    title=f"Hierarchy: {' → '.join(cols_to_use)}",
                    color=val_col,
                    color_continuous_scale="Viridis"
                )
            else:
                icicle_df = pdf.dropna(subset=cols_to_use)
                icicle_df['_count'] = 1
                fig = px.icicle(
                    icicle_df,
                    path=cols_to_use,
                    values='_count',
                    title=f"Hierarchy: {' → '.join(cols_to_use)}"
                )
            
            fig.update_traces(
                textinfo="label+percent parent",
                textfont=dict(size=12),       # ← Improved legibility
                marker=dict(line=dict(width=2, color="white")),
            )
            fig.update_layout(
                template="plotly_dark",
                height=460,                   # ← taller so text breathes
                margin=dict(l=10, r=10, t=10, b=10)
            )
            score = 55 + (len(cols_to_use) * 10)
            charts.append(ChartData(
                chart_id=f"icicle_{'_'.join(cols_to_use)}",
                chart_type="icicle",
                title=f"Icicle Chart",
                description=f"Hierarchical breakdown across {len(cols_to_use)} levels",
                plotly_json=json.loads(fig.update_layout(**CHART_LAYOUT_BASE).to_json()),
                columns_used=cols_to_use + ([val_col] if val_col else []),
                priority_score=score,
                insight_reason=f"{len(cols_to_use)}-level hierarchy",
                interest_level=get_interest_level(score)
            ))
        except:
            pass
    
    # ============== PHASE 3: ADVANCED CHARTS ==============
    
    # 3.1 PARALLEL COORDINATES: Multi-numeric exploration (enhanced)
    if is_chart_allowed("parallel_coordinates") and len(numeric_cols) >= 4 and len(charts) < max_charts:
        dims = numeric_cols[:6]  # Limit to 6 dimensions
        color_col = best_cat if best_cat else (categorical_cols[0] if categorical_cols else None)
        try:
            cols_to_plot = numeric_cols[:5]
            dimensions = []
            for col in cols_to_plot:
                col_data = pdf[col].dropna()
                if len(col_data) == 0: continue
                dimensions.append(dict(
                    range=[col_data.min(), col_data.max()],
                    label=col,                   # ← axis name shown clearly
                    values=pdf[col],
                    tickformat=".0f",
                ))

            if color_col and pdf[color_col].nunique() <= 8:
                codes, _ = pd.factorize(pdf[color_col])
                line_config = dict(
                    color=codes,
                    colorscale="Viridis",
                    showscale=True,
                    colorbar=dict(title=color_col),
                )
            else:
                line_config = dict(color="#6366f1")

            fig = go.Figure(data=go.Parcoords(
                line=line_config,
                dimensions=dimensions,
            ))

            fig.update_layout(
                title=dict(text=""),
                template="plotly_dark",
                font=dict(size=12, color="#94a3b8"),
                margin=dict(l=80, r=80, t=40, b=40),   # ← extra margin for axis labels
                height=420,
            )
            score = 50 + len(cols_to_plot) * 8
            charts.append(ChartData(
                chart_id="parallel_coords",
                chart_type="parallel_coordinates",
                title="Parallel Coordinates",
                description=f"Compare {len(cols_to_plot)} numeric dimensions simultaneously",
                plotly_json=json.loads(fig.update_layout(**CHART_LAYOUT_BASE).to_json()),
                columns_used=cols_to_plot,
                priority_score=score,
                insight_reason=f"{len(cols_to_plot)} numeric columns suitable for parallel comparison",
                interest_level=get_interest_level(score)
            ))
        except:
            pass
    
    # 3.2 SANKEY DIAGRAM: Flow between categorical columns
    if is_chart_allowed("sankey") and len(categorical_cols) >= 2 and len(charts) < max_charts:
        # Find best source-target pair with good cardinality
        source_col = None
        target_col = None
        best_flow_score = 0
        
        for i, col1 in enumerate(categorical_cols[:4]):
            for col2 in categorical_cols[i+1:4]:
                n1, n2 = pdf[col1].nunique(), pdf[col2].nunique()
                if 2 <= n1 <= 10 and 2 <= n2 <= 10:
                    # Score based on cardinality balance and distinct values
                    flow_score = min(n1, n2) * max(n1, n2) / (abs(n1 - n2) + 1)
                    if flow_score > best_flow_score:
                        best_flow_score = flow_score
                        source_col, target_col = col1, col2
        if source_col and target_col:
            try:
                # Create source-target-value aggregation
                flow_df = pdf.groupby([source_col, target_col]).size().reset_index(name='count')
                
                # HUMAN-READABLE LABELS (Fix 6)
                src_values = sorted(pdf[source_col].dropna().unique().tolist())
                tgt_values = sorted(pdf[target_col].dropna().unique().tolist())
                
                src_labels = [f"{source_col}: {v}" for v in src_values]
                tgt_labels = [f"{target_col}: {v}" for v in tgt_values]
                all_labels = src_labels + tgt_labels
                
                src_idx_map = {v: i for i, v in enumerate(src_values)}
                tgt_idx_map = {v: i + len(src_values) for i, v in enumerate(tgt_values)}
                
                # Generate link indices
                source_idx = [src_idx_map[s] for s in flow_df[source_col]]
                target_idx = [tgt_idx_map[t] for t in flow_df[target_col]]
                
                fig = go.Figure(go.Sankey(
                    node=dict(
                        pad=18,
                        thickness=22,
                        line=dict(color="rgba(255,255,255,0.3)", width=0.5),
                        label=all_labels,
                        color="#6366f1",
                        hovertemplate='%{label}<br>Total: %{value}<extra></extra>'
                    ),
                    link=dict(
                        source=source_idx,
                        target=target_idx,
                        value=flow_df['count'].tolist(),
                        color='rgba(139, 92, 246, 0.4)',
                        hovertemplate='%{source.label} → %{target.label}<br>Count: %{value}<extra></extra>'
                    )
                ))
                fig.update_layout(
                    title=f"Flow: {source_col} → {target_col}",
                    template="plotly_dark",
                    font=dict(size=12, color="#94a3b8"),
                    margin=dict(l=10, r=10, t=10, b=10),
                    height=420,
                )
                score = 70 + min(best_flow_score, 20)  # High boost for Sankey
                charts.append(ChartData(
                    chart_id=f"sankey_{source_col}_{target_col}",
                    chart_type="sankey",
                    title=f"Flow Diagram",
                    description=f"Flow relationships between {source_col} and {target_col}",
                    plotly_json=json.loads(fig.update_layout(**CHART_LAYOUT_BASE).to_json()),
                    columns_used=[source_col, target_col],
                    priority_score=score,
                    insight_reason=f"Flow structure detected between {source_col} and {target_col}",
                    interest_level=get_interest_level(score)
                ))
            except:
                pass
    
    # 3.3 RADAR / POLAR CHART: Multi-attribute comparison
    if is_chart_allowed("radar") and len(numeric_cols) >= 4 and len(numeric_cols) <= 8 and len(charts) < max_charts:
        radar_cols = numeric_cols[:8]
        try:
            # Normalize values for radar (0-1 scale)
            radar_df = pdf[radar_cols].copy()
            for col in radar_cols:
                min_val, max_val = radar_df[col].min(), radar_df[col].max()
                if max_val > min_val:
                    radar_df[col] = (radar_df[col] - min_val) / (max_val - min_val)
            
            # Use median values for the radar
            median_values = radar_df.median().tolist()
            median_values.append(median_values[0])  # Close the polygon
            radar_cols_closed = radar_cols + [radar_cols[0]]
            
            fig = go.Figure()
            fig.add_trace(go.Scatterpolar(
                r=median_values,
                theta=radar_cols_closed,
                fill='toself',
                fillcolor='rgba(99, 102, 241, 0.3)',
                line=dict(color='#6366f1', width=2),
                name='Median'
            ))
            fig.update_layout(
                title=f"Attribute Comparison Radar ({len(radar_cols)} dimensions)",
                template="plotly_dark",
                polar=dict(
                    radialaxis=dict(visible=True, range=[0, 1]),
                    bgcolor='rgba(0,0,0,0)'
                )
            )
            score = 55 + len(radar_cols) * 5
            charts.append(ChartData(
                chart_id="radar_attributes",
                chart_type="radar",
                title=f"Attribute Radar",
                description=f"Multi-attribute comparison across {len(radar_cols)} dimensions",
                plotly_json=json.loads(fig.update_layout(**CHART_LAYOUT_BASE).to_json()),
                columns_used=radar_cols,
                priority_score=score,
                insight_reason=f"{len(radar_cols)} numeric attributes suitable for radar comparison",
                interest_level=get_interest_level(score)
            ))
        except:
            pass
    
    # 3.4 WATERFALL CHART: For numeric with positive/negative values
    if is_chart_allowed("waterfall"):
        waterfall_col = None
        waterfall_score = 0
        for col in numeric_cols:
            try:
                values = pdf[col].dropna()
                has_positive = (values > 0).any()
                has_negative = (values < 0).any()
                if has_positive and has_negative:
                    # Good candidate - has both positive and negative
                    ratio = min(abs(values[values > 0].sum()), abs(values[values < 0].sum())) / max(abs(values.sum()), 1)
                    if ratio > 0.1:  # At least 10% contribution from both sides
                        col_score = 80 + ratio * 20
                        if col_score > waterfall_score:
                            waterfall_score = col_score
                            waterfall_col = col
            except:
                pass
    
    if is_chart_allowed("waterfall") and waterfall_col and len(charts) < max_charts:
        try:
            # Create waterfall from top values
            cat_col = best_cat if best_cat else (categorical_cols[0] if categorical_cols else None)
            if cat_col and pdf[cat_col].nunique() <= 10:
                waterfall_data = pdf.groupby(cat_col)[waterfall_col].sum().sort_values(ascending=False).head(10)
                
                fig = go.Figure(go.Waterfall(
                    x=waterfall_data.index.tolist(),
                    y=waterfall_data.values.tolist(),
                    connector=dict(line=dict(color='rgba(99, 102, 241, 0.5)', width=1.5, dash='dash')),
                    increasing=dict(marker=dict(color='#10b981', line=dict(color='#059669', width=1))),
                    decreasing=dict(marker=dict(color='#ef4444', line=dict(color='#dc2626', width=1))),
                    totals=dict(marker=dict(color='#6366f1', line=dict(color='#4f46e5', width=1))),
                    textposition="outside",
                    hovertemplate='%{x}<br>%{y:+,.0f}<extra></extra>'
                ))
                net_value = waterfall_data.sum()
                fig.update_layout(
                    title=f"Waterfall: {waterfall_col} by {cat_col}",
                    template="plotly_dark",
                    annotations=[dict(
                        x=1, y=1.1, xref="paper", yref="paper",
                        text=f"Net: {net_value:+,.0f}",
                        showarrow=False,
                        font=dict(size=14, color='#10b981' if net_value >= 0 else '#ef4444')
                    )]
                )
                charts.append(ChartData(
                    chart_id=f"waterfall_{waterfall_col}",
                    chart_type="waterfall",
                    title=f"Waterfall Chart",
                    description=f"Cumulative {waterfall_col} changes by {cat_col}",
                    plotly_json=json.loads(fig.update_layout(**CHART_LAYOUT_BASE).to_json()),
                    columns_used=[waterfall_col, cat_col],
                    priority_score=waterfall_score,
                    insight_reason="Clear positive/negative value pattern detected",
                    interest_level=get_interest_level(waterfall_score)
                ))
        except:
            pass
    
    # 3.5 CORRELATION HEATMAP (enhanced with priority)
    if is_chart_allowed("heatmap") and len(numeric_cols) >= 3 and len(charts) < max_charts:
        try:
            corr_matrix = pdf[numeric_cols].corr()
            fig = go.Figure(data=go.Heatmap(
                z=corr_matrix.values,
                x=corr_matrix.columns.tolist(),
                y=corr_matrix.columns.tolist(),
                colorscale='RdBu_r',
                zmin=-1, zmax=1,
                text=corr_matrix.round(2).values,
                texttemplate="%{text}",
                textfont={"size": 10}
            ))
            fig.update_layout(
                title="Correlation Matrix",
                template="plotly_dark"
            )
            # Score based on how many strong correlations exist
            strong_corr_count = ((corr_matrix.abs() > 0.5) & (corr_matrix.abs() < 1)).sum().sum() / 2
            score = 50 + min(strong_corr_count * 5, 30)
            charts.append(ChartData(
                chart_id="correlation_heatmap",
                chart_type="heatmap",
                title="Correlation Matrix",
                description="Pairwise correlations between numeric columns",
                plotly_json=json.loads(fig.update_layout(**CHART_LAYOUT_BASE).to_json()),
                columns_used=numeric_cols,
                priority_score=score,
                insight_reason=f"{int(strong_corr_count)} strong correlations found" if strong_corr_count > 0 else "Overview of variable relationships",
                interest_level=get_interest_level(score)
            ))
        except:
            pass
    
    charts.sort(key=lambda c: c.priority_score, reverse=True)
    
    # Apply focus filter
    if focus == "high_insight":
        charts = [c for c in charts if c.priority_score >= 70]
    elif focus == "smart_pick":
         charts = [c for c in charts if c.interest_level in ["high", "recommended"]]

    return VizResponse(
        session_id=session_id,
        charts=charts[:max_charts],
        total_generated=len(charts)
    )


class ChatMessage(BaseModel):
    role: str  # "user" | "assistant"
    content: str
    chart_data: Optional[dict] = None

class ChatRequest(BaseModel):
    question: str

class ChatResponse(BaseModel):
    session_id: str
    answer: str
    chart_type: Optional[str] = None
    chart_data: Optional[dict] = None
    sql_equivalent: Optional[str] = None

@app.post("/chat/{session_id}", response_model=ChatResponse)
def chat_with_data(session_id: str, request: ChatRequest):
    """
    Natural language interface to query data.
    This is the 'Talk to Your Data' feature - converts questions to analysis.
    """
    try:
        filename, df = load_session(session_id)
    except FileNotFoundError:
        raise HTTPException(status_code=404, detail="Session not found")
    
    question = request.question.lower()
    
    # ── Smart column classification ──
    _classifier = ColumnClassifier()
    _profile = _classifier.classify(df)
    numeric_cols = _profile.numericals
    categorical_cols = _profile.categoricals

    # Use Gemini for intent detection
    if client:
        try:
            # Note: We use the new client structure
            response = client.models.generate_content(
                model='gemini-2.0-flash',
                contents=f"""
                You are a helpful chart refinement assistant. The user is asking to change how visualizations are displayed.

                Current columns: {', '.join(numeric_cols)} (numeric), {', '.join(categorical_cols)} (categorical)
                
                User said: "{question}"

                Parse the request and return ONLY valid JSON with these keys:
                {{
                    "intent": "chart_refinement" | "none" | "other",
                    "params": {{
                        "groupby": "ColumnName" | null,
                        "chart_types": ["bar", "pie", ...] | null,
                        "focus": "high_insight" | "smart_pick" | "all" | null,
                        "exclude_types": ["pie", ...] | null,
                        "specific_request": "description if unclear"
                    }},
                    "reply": "short friendly confirmation message to show user"
                }}

                If no chart-related request, return intent: "none" and reply with normal chat answer to the question based on data if possible (though you don't have row data here, just schema). If it's a general data question, just answer generally or ask them to be specific.
                """
            )
            result = json.loads(response.text.strip().replace('```json', '').replace('```', ''))
            
            intent = result.get("intent")
            params = result.get("params", {})
            reply = result.get("reply", "Done.")
            
            chart_response_data = None
            
            if intent == "chart_refinement":
                # Call generate_visualizations with params
                # Note: We need to adapt params to match function signature
                viz_response = generate_visualizations(
                    session_id=session_id,
                    max_charts=10,
                    groupby=params.get("groupby"),
                    chart_types=params.get("chart_types"),
                    focus=params.get("focus"),
                    exclude_types=params.get("exclude_types")
                )
                chart_response_data = viz_response.dict()
                answer = reply
                # Track refinement history
                if not session_id.isdigit():
                    try:
                        db.add_history(session_id, "refine", params, f"Chat refinement: {question[:80]}")
                    except Exception:
                        pass
            else:
                answer = reply

            return ChatResponse(
                session_id=session_id,
                answer=answer,
                chart_data=chart_response_data
            )

        except Exception as e:
            print(f"Gemini Error: {e}")
            # Fallback to simple logic or error message
            pass

    # Fallback legacy logic if Gemini fails or no key
    answer = "I'm listening, but my AI brain (Gemini) isn't connected or had an error. I can't refine charts yet."
    return ChatResponse(session_id=session_id, answer=answer)
    
    # Default: describe the dataset


# ============== PHASE 5: MODELING & REPORTING ==============
# NOTE: sklearn imports are deferred to function level to avoid slow startup

class ModelConfig(BaseModel):
    target_column: str
    goal: str  # "predict" | "classify" | "segment"
    features: Optional[list[str]] = None  # If None, use all numeric

class ModelResult(BaseModel):
    model_name: str
    model_type: str  # "regression" | "classification"
    accuracy_or_r2: float
    secondary_metric: float  # MSE for regression, F1 for classification
    feature_importance: dict[str, float]

class ModelingResponse(BaseModel):
    session_id: str
    target_column: str
    goal: str
    best_model: str
    models: list[ModelResult]
    prediction_sample: list[dict]

@app.post("/model/{session_id}", response_model=ModelingResponse)
def train_models(session_id: str, config: ModelConfig):
    """
    AutoML Pipeline: Train and compare multiple models.
    This is the 'Smart Modeling' feature - what a Data Scientist does for predictions.
    """
    # Deferred imports to avoid slow startup on Railway
    from sklearn.model_selection import train_test_split
    from sklearn.preprocessing import LabelEncoder, StandardScaler
    from sklearn.linear_model import LinearRegression, LogisticRegression
    from sklearn.tree import DecisionTreeClassifier, DecisionTreeRegressor
    from sklearn.ensemble import RandomForestClassifier, RandomForestRegressor
    from sklearn.metrics import accuracy_score, mean_squared_error, r2_score, f1_score
    import numpy as np
    
    try:
        filename, df = load_session(session_id)
    except FileNotFoundError:
        raise HTTPException(status_code=404, detail="Session not found")
    
    if config.target_column not in df.columns:
        raise HTTPException(status_code=400, detail=f"Target column '{config.target_column}' not found")
    
    # Prepare features
    numeric_cols = [col for col in df.columns if df[col].dtype in [pl.Float64, pl.Float32, pl.Int64, pl.Int32, pl.Int16, pl.Int8]]
    
    if config.features:
        feature_cols = [f for f in config.features if f in numeric_cols and f != config.target_column]
    else:
        feature_cols = [c for c in numeric_cols if c != config.target_column]
    
    if len(feature_cols) < 1:
        raise HTTPException(status_code=400, detail="Need at least 1 numeric feature column")
    
    # Prepare data
    df_clean = df.select(feature_cols + [config.target_column]).drop_nulls()
    if len(df_clean) < 20:
        raise HTTPException(status_code=400, detail="Need at least 20 rows after removing nulls")
    
    X = df_clean.select(feature_cols).to_numpy()
    y = df_clean[config.target_column].to_numpy()
    
    # Determine if classification or regression
    target_dtype = df[config.target_column].dtype
    unique_values = df[config.target_column].n_unique()
    
    is_classification = (
        config.goal == "classify" or 
        target_dtype == pl.Utf8 or 
        (unique_values <= 10 and target_dtype in [pl.Int64, pl.Int32, pl.Int16, pl.Int8])
    )
    
    # Encode target if classification with strings
    le = None
    if target_dtype == pl.Utf8:
        le = LabelEncoder()
        y = le.fit_transform(y)
    
    # Scale features
    scaler = StandardScaler()
    X_scaled = scaler.fit_transform(X)
    
    # Split data
    X_train, X_test, y_train, y_test = train_test_split(X_scaled, y, test_size=0.2, random_state=42)
    
    results = []
    
    if is_classification:
        # Classification models
        models = [
            ("Logistic Regression", LogisticRegression(max_iter=1000)),
            ("Decision Tree", DecisionTreeClassifier(max_depth=5)),
            ("Random Forest", RandomForestClassifier(n_estimators=50, max_depth=5))
        ]
        
        for name, model in models:
            try:
                model.fit(X_train, y_train)
                y_pred = model.predict(X_test)
                acc = accuracy_score(y_test, y_pred)
                f1 = f1_score(y_test, y_pred, average='weighted')
                
                # Feature importance
                if hasattr(model, 'feature_importances_'):
                    importance = dict(zip(feature_cols, [round(float(x), 4) for x in model.feature_importances_]))
                elif hasattr(model, 'coef_'):
                    coef = model.coef_[0] if len(model.coef_.shape) > 1 else model.coef_
                    importance = dict(zip(feature_cols, [round(abs(float(x)), 4) for x in coef]))
                else:
                    importance = {}
                
                results.append(ModelResult(
                    model_name=name,
                    model_type="classification",
                    accuracy_or_r2=round(acc, 4),
                    secondary_metric=round(f1, 4),
                    feature_importance=importance
                ))
            except Exception as e:
                pass
    else:
        # Regression models
        models = [
            ("Linear Regression", LinearRegression()),
            ("Decision Tree", DecisionTreeRegressor(max_depth=5)),
            ("Random Forest", RandomForestRegressor(n_estimators=50, max_depth=5))
        ]
        
        for name, model in models:
            try:
                model.fit(X_train, y_train)
                y_pred = model.predict(X_test)
                r2 = r2_score(y_test, y_pred)
                mse = mean_squared_error(y_test, y_pred)
                
                # Feature importance
                if hasattr(model, 'feature_importances_'):
                    importance = dict(zip(feature_cols, [round(float(x), 4) for x in model.feature_importances_]))
                elif hasattr(model, 'coef_'):
                    importance = dict(zip(feature_cols, [round(abs(float(x)), 4) for x in model.coef_]))
                else:
                    importance = {}
                
                results.append(ModelResult(
                    model_name=name,
                    model_type="regression",
                    accuracy_or_r2=round(r2, 4),
                    secondary_metric=round(mse, 4),
                    feature_importance=importance
                ))
            except Exception as e:
                pass
    
    if not results:
        raise HTTPException(status_code=500, detail="Failed to train any models")
    
    # Find best model
    best = max(results, key=lambda x: x.accuracy_or_r2)
    
    # Generate sample predictions
    sample_indices = list(range(min(5, len(X_test))))
    predictions = []
    for i in sample_indices:
        pred_row = {"actual": float(y_test[i]) if not le else le.inverse_transform([int(y_test[i])])[0]}
        for name, model in models:
            try:
                pred = model.predict(X_test[i:i+1])[0]
                pred_row[name] = float(pred) if not le else le.inverse_transform([int(pred)])[0]
            except:
                pass
        predictions.append(pred_row)
    
    return ModelingResponse(
        session_id=session_id,
        target_column=config.target_column,
        goal=config.goal,
        best_model=best.model_name,
        models=results,
        prediction_sample=predictions
    )


class ReportSection(BaseModel):
    title: str
    content: str
    chart_type: Optional[str] = None
    chart_data: Optional[dict] = None

class ReportResponse(BaseModel):
    session_id: str
    title: str
    generated_at: str
    sections: list[ReportSection]

@app.get("/report/{session_id}", response_model=ReportResponse)
def generate_report(session_id: str):
    """
    Generate a comprehensive analysis report.
    This is the 'Communication' step - presenting findings to stakeholders.
    """
    from datetime import datetime
    
    try:
        filename, df = load_session(session_id)
    except FileNotFoundError:
        raise HTTPException(status_code=404, detail="Session not found")
    
    sections = []
    
    # Section 1: Dataset Overview
    numeric_cols = [col for col in df.columns if df[col].dtype in [pl.Float64, pl.Float32, pl.Int64, pl.Int32, pl.Int16, pl.Int8]]
    categorical_cols = [col for col in df.columns if df[col].dtype == pl.Utf8]
    
    sections.append(ReportSection(
        title="Dataset Overview",
        content=f"This report analyzes '{filename}' containing {len(df):,} records across {len(df.columns)} columns. The dataset includes {len(numeric_cols)} numeric columns and {len(categorical_cols)} categorical columns.",
    ))
    
    # Section 2: Data Quality
    missing_total = sum(df[col].null_count() for col in df.columns)
    duplicate_count = len(df) - df.unique().height
    quality_score = "Good" if missing_total < len(df) * 0.1 else ("Fair" if missing_total < len(df) * 0.3 else "Needs Attention")
    
    sections.append(ReportSection(
        title="Data Quality Assessment",
        content=f"Data Quality: {quality_score}. Missing values: {missing_total:,} ({round(missing_total/len(df)/len(df.columns)*100, 1)}% of cells). Duplicate rows: {duplicate_count}.",
    ))
    
    # Section 3: Key Statistics
    if numeric_cols:
        stats_text = "Key Metrics:\n"
        for col in numeric_cols[:3]:
            col_data = df[col].drop_nulls()
            if len(col_data) > 0:
                stats_text += f"• {col}: Mean = {col_data.mean():.2f}, Median = {col_data.median():.2f}, Range = [{col_data.min():.2f}, {col_data.max():.2f}]\n"
        sections.append(ReportSection(
            title="Statistical Summary",
            content=stats_text.strip(),
        ))
    
    # Section 4: Strategic Brief (V2 Alignment)
    try:
        # Pull from the new synthesized intelligence layer
        insights_data = get_insights(session_id)
        sections.append(ReportSection(
            title="Strategic Executive Intelligence",
            content=insights_data.executive_summary,
        ))
        
        brief_text = "Key Strategic Findings:\n"
        for ins in insights_data.strategic_brief[:5]:
            brief_text += f"• {ins.title}: {ins.description}\n"
            
        sections.append(ReportSection(
            title="Deep Insights & Findings",
            content=brief_text.strip(),
        ))
        insight_count = len(insights_data.strategic_brief)
    except Exception as e:
        print(f"Report Insight Error: {e}")
        # Legacy fallback if insights engine fails
        sections.append(ReportSection(
            title="Statistical Summary",
            content="Historical data patterns detected across majority segments.",
        ))
    
    # Section 5: Recommendations
    recommendations = "Recommended Actions:\n"
    recommendations += "1. Address any missing values before modeling\n"
    recommendations += "2. Investigate correlations for potential causal relationships\n"
    recommendations += "3. Consider segmenting analysis by dominant categories\n"
    
    sections.append(ReportSection(
        title="Recommendations",
        content=recommendations.strip(),
    ))
    
    return ReportResponse(
        session_id=session_id,
        title=f"Analysis Report: {filename}",
        generated_at=datetime.now().isoformat(),
        sections=sections
    )

@app.get("/export-excel/{session_id}")
def export_excel_report(session_id: str):
    """
    Generate a comprehensive multi-sheet Excel report.
    Sheet 1: Cleaned Data
    Sheet 2: Executive Summary & Insights
    Sheet 3: Visualizations (as images)
    """
    try:
        # 1. Get Data
        filename, df = load_session(session_id)
        df_pandas = df.to_pandas()
        
        # 2. Get Insights
        insights_data = get_insights(session_id)
        insights = insights_data.strategic_brief
        
        # 3. Get Chart Titles from Engine
        recommender = SmartChartRecommender()
        chart_titles = recommender.get_chart_titles(df)
        
        # 4. Create Workbook
        wb = Workbook()
        
        # ========================
        # Sheet 1: Cleaned Data
        # ========================
        ws_data = wb.active
        ws_data.title = "Cleaned Data"
        
        # Write dataframe
        for r_idx, row in enumerate(dataframe_to_rows(df_pandas, index=False, header=True), 1):
            ws_data.append(row)
            
        # Write header with style
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
        for cell in ws_data[1]:
            cell.font = header_font
            cell.fill = header_fill
            
        # Freeze top row
        ws_data.freeze_panes = 'A2'

        # ========================
        # Sheet 2: Insights
        # ========================
        ws_insights = wb.create_sheet("Insights")
        ws_insights.column_dimensions['A'].width = 110
        
        # Executive Summary
        ws_insights.append(["EXECUTIVE SUMMARY"])
        ws_insights['A1'].font = Font(size=14, bold=True, color="4F46E5")
        ws_insights.append([insights_data.executive_summary])
        ws_insights['A2'].alignment = Alignment(wrap_text=True)
        ws_insights.append([""])
        
        critical_insights = [ins for ins in insights if ins.impact in ("Critical", "High", "🔴 Critical")]
        if not critical_insights:
            ws_insights.append(["No Critical/High-impact strategic risks detected."])
            ws_insights[f'A{ws_insights.max_row}'].font = Font(italic=True)
        else:
            ws_insights.append(["PRIORITY STRATEGIC RISKS"])
            ws_insights[f'A{ws_insights.max_row}'].font = Font(bold=True)
            for ins in critical_insights:
                ws_insights.append([f"• [{ins.impact}] {ins.title}"])
        
        ws_insights.append([""])
        ws_insights.append(["KEY ANALYTICAL FINDINGS"])
        ws_insights[f'A{ws_insights.max_row}'].font = Font(size=12, bold=True)
        
        # Group insights by impact
        from collections import defaultdict
        by_impact = defaultdict(list)
        for ins in insights:
            # Normalize impact string for grouping
            imp = ins.impact.replace("🔴 ", "").replace("🟠 ", "").replace("🟢 ", "").title()
            by_impact[imp].append(ins)
            
        # Define colors
        impact_colors = {
            "Critical": "FF0000",
            "High": "FF0000",
            "Important": "FFA500",
            "Medium": "FFA500",
            "Minor": "008000",
            "Low": "008000"
        }
        
        # Sort impacts by severity
        order = ["Critical", "High", "Important", "Medium", "Minor", "Low"]
        for impact_key in order:
            if impact_key not in by_impact:
                continue
            
            # Write impact header
            ws_insights.append([f"--- {impact_key} Insights ---"])
            for cell in ws_insights[ws_insights.max_row]:
                cell.font = Font(bold=True, color=impact_colors.get(impact_key, "000000"))
            
            for ins in by_impact[impact_key]:
                ws_insights.append([f"• {ins.title}"])
                ws_insights[f'A{ws_insights.max_row}'].font = Font(bold=True)
                ws_insights.append([ins.description.replace("**", "")])
                ws_insights[f'A{ws_insights.max_row}'].alignment = Alignment(wrap_text=True)
                ws_insights.append([f"Recommendation: {ins.recommendation}"])
                ws_insights[f'A{ws_insights.max_row}'].font = Font(italic=True)
                ws_insights.append([""])
                
        # Recommendations Summary
        ws_insights.append([""])
        ws_insights.append(["ACTIONABLE RECOMMENDATIONS"])
        ws_insights[f'A{ws_insights.max_row}'].font = Font(size=12, bold=True)
        recs = [ins.recommendation for ins in insights if ins.recommendation and "No specific" not in ins.recommendation]
        for rec in recs:
            ws_insights.append([f"• {rec}"])

        # ========================
        # Sheet 3: Visualizations
        # ========================
        ws_viz = wb.create_sheet("Visualizations")
        ws_viz.column_dimensions['A'].width = 100
        
        # List chart titles here
        ws_viz.append(["DASHBOARD VISUALIZATIONS"])
        ws_viz['A1'].font = Font(size=14, bold=True, color="4F46E5")
        ws_viz.append([""])
        
        # We reuse _internal_gen_viz for actual images if we wanted, 
        # but the user patch only lists titles for now.
        for chart_name, title in chart_titles.items():
            ws_viz.append([title])
            ws_viz[f'A{ws_viz.max_row}'].font = Font(bold=True)
            ws_viz.append(["[Image Asset Placeholder]"])
            ws_viz.append([""])

        # Save to buffer
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        export_filename = f"InsightStream_Report_{session_id[:8]}.xlsx"
        
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={export_filename}"}
        )

    except Exception as e:
         print(f"Export Error: {e}")
         raise HTTPException(status_code=500, detail=f"Export failed: {str(e)}")

    except Exception as e:
         print(f"Export Error: {e}")
         raise HTTPException(status_code=500, detail=f"Export failed: {str(e)}")

def is_port_in_use(port: int) -> bool:
    """Check if a port is already in use."""
    with socket.socket(socket.AF_INET, socket.SOCK_STREAM) as s:
        try:
            s.bind(("0.0.0.0", port))
            return False
        except OSError:
            return True

def find_free_port(start_port: int) -> int:
    """Find next available port from start_port."""
    port = start_port
    while port < start_port + 10:
        if not is_port_in_use(port):
            return port
        port += 1
    raise RuntimeError(f"No free ports found between {start_port} and {start_port + 10}")

@app.post("/dev/load-csv/{session_id}")
async def dev_load_csv(session_id: str, filepath: str):
    """
    Dev-only endpoint. Loads a CSV from the server's filesystem into the session store.
    """
    try:
        df = pd.read_csv(filepath)
        # Convert to polars and save to session dir
        pl_df = pl.from_pandas(df)
        session_path = SESSION_DIR / session_id
        session_path.mkdir(parents=True, exist_ok=True)
        pl_df.write_parquet(session_path / "data.parquet")
        
        # Metadata
        metadata = {
            "original_filename": os.path.basename(filepath),
            "rows": len(df),
            "columns": list(df.columns)
        }
        with open(session_path / "metadata.json", "w") as f:
            json.dump(metadata, f)
            
        return {"status": "ok", "rows": len(df), "columns": list(df.columns)}
    except Exception as exc:
        raise HTTPException(status_code=400, detail=str(exc))

if __name__ == '__main__':
    import uvicorn
    
    preferred_port = int(os.environ.get('PORT', 8000))
    
    if is_port_in_use(preferred_port):
        print(f"Port {preferred_port} is in use. Finding alternative...")
        try:
            port = find_free_port(preferred_port + 1)
            print(f"Using port {port} instead.")
            print(f"Update frontend API_BASE to http://localhost:{port}")
        except RuntimeError as e:
            print(f"Error: {e}")
            sys.exit(1)
    else:
        port = preferred_port
        
    print(f"Starting InsightStream on port {port}...")
    # Note: Uvicorn doesn't have a direct body size limit parameter
    # The limit is controlled by the ASGI server (usually unlimited for streaming uploads)
    # If needed, configure via environment: UVICORN_LIMIT_MAX_REQUESTS
    uvicorn.run(
        app, 
        host='0.0.0.0', 
        port=port, 
        proxy_headers=True, 
        forwarded_allow_ips="*",
        timeout_keep_alive=300  # 5 minutes for large uploads
    )
